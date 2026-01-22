#!/usr/bin/env python3
#
# Copyright 2022 CNApy organization
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
# -*- coding: utf-8 -*-

"""Enhanced Flux Response Analysis dialog.

Scan a target reaction flux from min..max on the x-axis.
For each fixed value, maximize a product reaction (y-axis) and record biomass.
Provides comprehensive visualization, results table, and export options.

Features:
- Auto-detect flux range via FVA
- Wild-type comparison
- 4-panel visualization
- Results table with optimal point highlight
- CSV/XLSX export
- Background thread execution
"""

import csv

import numpy as np
from qtpy.QtCore import Qt, QThread, Signal, Slot
from qtpy.QtWidgets import (
    QCheckBox,
    QComboBox,
    QDialog,
    QDoubleSpinBox,
    QFileDialog,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QSpinBox,
    QSplitter,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from cnapy.appdata import AppData

# Check for openpyxl availability for XLSX export
try:
    import openpyxl

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Check for matplotlib availability for plots
try:
    from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
    from matplotlib.figure import Figure

    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False


class FluxResponseWorkerThread(QThread):
    """Worker thread for running flux response analysis in background."""

    progress_update = Signal(int, int)  # (current, total)
    result_ready = Signal(object)  # dict with results
    error_occurred = Signal(str)

    def __init__(
        self,
        model,
        appdata: AppData,
        target_reaction: str,
        product_reaction: str,
        biomass_reaction: str,
        x_min: float,
        x_max: float,
        n_steps: int,
        use_scenario: bool,
    ):
        """Initialize the worker thread.

        Args:
            model: COBRApy model (copy)
            appdata: Application data for scenario loading
            target_reaction: Reaction ID to scan (x-axis)
            product_reaction: Reaction ID to maximize (y-axis)
            biomass_reaction: Reaction ID of biomass (for tracking)
            x_min: Minimum target flux
            x_max: Maximum target flux
            n_steps: Number of scanning steps
            use_scenario: Whether to apply scenario constraints
        """
        super().__init__()
        self.model = model
        self.appdata = appdata
        self.target_reaction = target_reaction
        self.product_reaction = product_reaction
        self.biomass_reaction = biomass_reaction
        self.x_min = x_min
        self.x_max = x_max
        self.n_steps = n_steps
        self.use_scenario = use_scenario
        self._cancel_requested = False

    def request_cancel(self):
        """Request cancellation of the analysis."""
        self._cancel_requested = True

    def run(self):
        """Execute the flux response analysis."""
        try:
            model = self.model

            # Load scenario if requested
            if self.use_scenario and self.appdata and self.appdata.project:
                try:
                    self.appdata.project.load_scenario_into_model(model)
                except Exception:
                    pass  # Best-effort scenario loading

            # Step 1: Wild-type FBA (maximize biomass)
            model.objective = model.reactions.get_by_id(self.biomass_reaction)
            model.objective_direction = "max"
            wt_sol = model.optimize()

            if wt_sol.status != "optimal":
                self.error_occurred.emit("Wild-type optimization failed. Check model constraints.")
                return

            wt_biomass = wt_sol.fluxes[self.biomass_reaction]
            wt_target = wt_sol.fluxes[self.target_reaction]
            wt_product = wt_sol.fluxes[self.product_reaction]

            # Step 2: Generate scan points
            scan_points = np.linspace(self.x_min, self.x_max, self.n_steps)
            scan_results = []

            target_rxn = model.reactions.get_by_id(self.target_reaction)
            product_rxn = model.reactions.get_by_id(self.product_reaction)
            biomass_rxn = model.reactions.get_by_id(self.biomass_reaction)

            # Set objective to product
            model.objective = product_rxn
            model.objective_direction = "max"

            self.progress_update.emit(0, self.n_steps)

            for idx, target_flux in enumerate(scan_points):
                if self._cancel_requested:
                    return

                with model:
                    # Fix target reaction
                    target_rxn.bounds = (target_flux, target_flux)

                    # Maximize product
                    sol = model.optimize()

                    if sol.status == "optimal":
                        scan_results.append(
                            {
                                "target_flux": target_flux,
                                "product_flux": sol.fluxes[self.product_reaction],
                                "biomass_flux": sol.fluxes[self.biomass_reaction],
                                "status": "optimal",
                            }
                        )
                    else:
                        scan_results.append(
                            {
                                "target_flux": target_flux,
                                "product_flux": np.nan,
                                "biomass_flux": np.nan,
                                "status": "infeasible",
                            }
                        )

                self.progress_update.emit(idx + 1, self.n_steps)

            # Step 3: Find optimal point (max product)
            optimal_data = [r for r in scan_results if r["status"] == "optimal"]
            if optimal_data:
                optimal_idx = np.nanargmax([r["product_flux"] for r in optimal_data])
                optimal_point = optimal_data[optimal_idx]
            else:
                optimal_point = None

            # Step 4: Calculate statistics
            feasible_results = [r for r in scan_results if r["status"] == "optimal"]
            product_values = [r["product_flux"] for r in feasible_results]
            biomass_values = [r["biomass_flux"] for r in feasible_results]

            stats = {}
            if product_values:
                stats["product_mean"] = np.mean(product_values)
                stats["product_std"] = np.std(product_values)
                stats["product_min"] = np.min(product_values)
                stats["product_max"] = np.max(product_values)
                stats["biomass_mean"] = np.mean(biomass_values)
                stats["biomass_std"] = np.std(biomass_values)
                stats["biomass_min"] = np.min(biomass_values)
                stats["biomass_max"] = np.max(biomass_values)

            self.result_ready.emit(
                {
                    "scan_results": scan_results,
                    "wt_biomass": wt_biomass,
                    "wt_target": wt_target,
                    "wt_product": wt_product,
                    "optimal_point": optimal_point,
                    "target_reaction": self.target_reaction,
                    "product_reaction": self.product_reaction,
                    "biomass_reaction": self.biomass_reaction,
                    "x_min": self.x_min,
                    "x_max": self.x_max,
                    "n_steps": self.n_steps,
                    "n_feasible": len(feasible_results),
                    "stats": stats,
                }
            )

        except Exception as e:
            import traceback

            self.error_occurred.emit(f"Flux response analysis failed: {str(e)}\n{traceback.format_exc()}")


class FluxResponseDialog(QDialog):
    """Enhanced dialog for flux response analysis."""

    def __init__(self, appdata: AppData, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Flux Response Analysis")
        self.setMinimumSize(1100, 750)

        self.appdata = appdata
        self.worker_thread: FluxResponseWorkerThread | None = None
        self.last_results: dict | None = None

        self._setup_ui()

    def _setup_ui(self):
        """Setup the dialog UI."""
        main_layout = QVBoxLayout()

        # Description
        desc_label = QLabel(
            "Flux Response Analysis scans a target reaction flux from min to max, "
            "maximizing a product reaction at each point. This helps identify optimal "
            "flux configurations and visualize the trade-off between target manipulation "
            "and product formation."
        )
        desc_label.setWordWrap(True)
        main_layout.addWidget(desc_label)

        # Parameters group
        params_group = QGroupBox("Analysis Parameters")
        params_layout = QVBoxLayout()

        # Target reaction
        target_layout = QHBoxLayout()
        target_layout.addWidget(QLabel("Target Reaction (x-axis):"))
        self.target_selector = QComboBox()
        self.target_selector.setMinimumWidth(350)
        self.target_selector.setEditable(True)
        self.target_selector.setToolTip("Reaction to scan from min to max. Its flux will be fixed at each scan point.")
        self._populate_reaction_selector(self.target_selector)
        target_layout.addWidget(self.target_selector)

        self.autodetect_btn = QPushButton("Auto-detect Range")
        self.autodetect_btn.setToolTip("Run FVA to find the feasible flux range for target reaction")
        self.autodetect_btn.clicked.connect(self._autodetect_range)
        target_layout.addWidget(self.autodetect_btn)
        target_layout.addStretch()
        params_layout.addLayout(target_layout)

        # Product reaction
        product_layout = QHBoxLayout()
        product_layout.addWidget(QLabel("Product Reaction (y-axis):"))
        self.product_selector = QComboBox()
        self.product_selector.setMinimumWidth(350)
        self.product_selector.setEditable(True)
        self.product_selector.setToolTip(
            "Reaction to maximize at each scan point. Typically an exchange reaction for product secretion."
        )
        self._populate_reaction_selector(self.product_selector, prioritize_exchange=True)
        product_layout.addWidget(self.product_selector)
        product_layout.addStretch()
        params_layout.addLayout(product_layout)

        # Biomass reaction
        biomass_layout = QHBoxLayout()
        biomass_layout.addWidget(QLabel("Biomass Reaction:"))
        self.biomass_selector = QComboBox()
        self.biomass_selector.setMinimumWidth(350)
        self.biomass_selector.setToolTip(
            "Biomass/objective reaction to track during analysis (not constrained by default)"
        )
        self._populate_reaction_selector(self.biomass_selector, prioritize_objective=True)
        biomass_layout.addWidget(self.biomass_selector)
        biomass_layout.addStretch()
        params_layout.addLayout(biomass_layout)

        # Range and steps
        range_layout = QHBoxLayout()
        range_layout.addWidget(QLabel("Min x:"))
        self.min_spin = QDoubleSpinBox()
        self.min_spin.setDecimals(4)
        self.min_spin.setRange(-1e9, 1e9)
        self.min_spin.setValue(-10.0)
        self.min_spin.setToolTip("Minimum flux value for target reaction")
        range_layout.addWidget(self.min_spin)

        range_layout.addWidget(QLabel("Max x:"))
        self.max_spin = QDoubleSpinBox()
        self.max_spin.setDecimals(4)
        self.max_spin.setRange(-1e9, 1e9)
        self.max_spin.setValue(10.0)
        self.max_spin.setToolTip("Maximum flux value for target reaction")
        range_layout.addWidget(self.max_spin)

        range_layout.addSpacing(20)

        range_layout.addWidget(QLabel("Steps:"))
        self.steps_spin = QSpinBox()
        self.steps_spin.setRange(5, 500)
        self.steps_spin.setValue(20)
        self.steps_spin.setToolTip("Number of scan points between min and max")
        range_layout.addWidget(self.steps_spin)

        range_layout.addStretch()
        params_layout.addLayout(range_layout)

        # Options and run button
        options_layout = QHBoxLayout()

        self.use_scenario = QCheckBox("Use scenario constraints")
        self.use_scenario.setChecked(True)
        self.use_scenario.setToolTip("Apply current scenario flux bounds to the model")
        options_layout.addWidget(self.use_scenario)

        options_layout.addStretch()

        # Run button
        self.run_btn = QPushButton("Run Analysis")
        self.run_btn.setStyleSheet("font-size: 14px; padding: 8px 16px; background-color: #4CAF50; color: white;")
        self.run_btn.clicked.connect(self._run_analysis)
        options_layout.addWidget(self.run_btn)

        # Cancel button
        self.cancel_btn = QPushButton("Cancel")
        self.cancel_btn.setEnabled(False)
        self.cancel_btn.clicked.connect(self._cancel_analysis)
        options_layout.addWidget(self.cancel_btn)

        params_layout.addLayout(options_layout)
        params_group.setLayout(params_layout)
        main_layout.addWidget(params_group)

        # Progress bar
        progress_layout = QHBoxLayout()
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        progress_layout.addWidget(self.progress_bar)

        self.status_label = QLabel("")
        self.status_label.setStyleSheet("color: gray;")
        progress_layout.addWidget(self.status_label)
        main_layout.addLayout(progress_layout)

        # Results splitter
        splitter = QSplitter(Qt.Horizontal)

        # Left side: Results table
        table_widget = QWidget()
        table_layout = QVBoxLayout()
        table_layout.setContentsMargins(0, 0, 0, 0)

        table_label = QLabel("Scan Results:")
        table_layout.addWidget(table_label)

        self.results_table = QTableWidget()
        self.results_table.setColumnCount(4)
        self.results_table.setHorizontalHeaderLabels(["Target Flux", "Product Flux", "Biomass", "Status"])
        self.results_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.results_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.results_table.setAlternatingRowColors(True)
        table_layout.addWidget(self.results_table)

        table_widget.setLayout(table_layout)
        splitter.addWidget(table_widget)

        # Right side: Plot
        if MATPLOTLIB_AVAILABLE:
            plot_widget = QWidget()
            plot_layout = QVBoxLayout()
            plot_layout.setContentsMargins(0, 0, 0, 0)

            self.figure = Figure(figsize=(8, 7), dpi=100)
            self.canvas = FigureCanvas(self.figure)
            plot_layout.addWidget(self.canvas)

            export_plot_btn = QPushButton("Export Plot...")
            export_plot_btn.clicked.connect(self._export_plot)
            plot_layout.addWidget(export_plot_btn)

            plot_widget.setLayout(plot_layout)
            splitter.addWidget(plot_widget)
        else:
            no_plot_label = QLabel("Install matplotlib for visualization")
            no_plot_label.setAlignment(Qt.AlignCenter)
            splitter.addWidget(no_plot_label)

        splitter.setSizes([350, 650])
        main_layout.addWidget(splitter)

        # Summary and export
        summary_layout = QHBoxLayout()

        self.summary_label = QLabel("No results yet. Click 'Run Analysis' to start.")
        self.summary_label.setWordWrap(True)
        summary_layout.addWidget(self.summary_label, stretch=3)

        summary_layout.addStretch()

        # Export buttons
        export_csv_btn = QPushButton("Export CSV")
        export_csv_btn.clicked.connect(self._export_csv)
        summary_layout.addWidget(export_csv_btn)

        export_xlsx_btn = QPushButton("Export XLSX")
        export_xlsx_btn.clicked.connect(self._export_xlsx)
        if not OPENPYXL_AVAILABLE:
            export_xlsx_btn.setEnabled(False)
            export_xlsx_btn.setToolTip("openpyxl not installed")
        summary_layout.addWidget(export_xlsx_btn)

        main_layout.addLayout(summary_layout)

        # Close button
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(self.accept)
        btn_layout.addWidget(close_btn)
        main_layout.addLayout(btn_layout)

        self.setLayout(main_layout)

    def _populate_reaction_selector(
        self, selector: QComboBox, prioritize_exchange: bool = False, prioritize_objective: bool = False
    ):
        """Populate a reaction selector combo box."""
        model = self.appdata.project.cobra_py_model

        objective_reactions = set()

        if prioritize_objective:
            # Try to find objective reactions first
            obj_found = False
            for rxn in model.reactions:
                if rxn.objective_coefficient != 0:
                    display_text = f"{rxn.id} - {rxn.name}" if rxn.name else rxn.id
                    selector.addItem(f"[Obj] {display_text}", rxn.id)
                    objective_reactions.add(rxn.id)
                    obj_found = True

            # If no objective reactions found, look for common biomass patterns
            if not obj_found:
                biomass_patterns = ["biomass", "growth", "bio_", "bof_"]
                for rxn in model.reactions:
                    rxn_id_lower = rxn.id.lower()
                    rxn_name_lower = (rxn.name or "").lower()
                    for pattern in biomass_patterns:
                        if pattern in rxn_id_lower or pattern in rxn_name_lower:
                            display_text = f"{rxn.id} - {rxn.name}" if rxn.name else rxn.id
                            selector.addItem(f"[Obj] {display_text}", rxn.id)
                            objective_reactions.add(rxn.id)
                            break

        if prioritize_exchange:
            # Add exchange reactions first
            exchange_rxns = [r for r in model.reactions if r.id.startswith("EX_")]
            for rxn in sorted(exchange_rxns, key=lambda r: r.id):
                display_text = f"{rxn.id} - {rxn.name}" if rxn.name else rxn.id
                selector.addItem(f"[Exchange] {display_text}", rxn.id)

        # Add all other reactions
        for rxn in sorted(model.reactions, key=lambda r: r.id):
            is_objective = rxn.id in objective_reactions or rxn.objective_coefficient != 0
            is_exchange = rxn.id.startswith("EX_")

            if (prioritize_objective and is_objective) or (prioritize_exchange and is_exchange):
                continue  # Already added

            display_text = f"{rxn.id} - {rxn.name}" if rxn.name else rxn.id
            selector.addItem(display_text, rxn.id)

    def _get_selected_reaction(self, selector: QComboBox) -> str:
        """Get the selected reaction ID from a combo box."""
        # First check if there's user data
        data = selector.currentData()
        if data:
            return data

        # Otherwise parse the text
        text = selector.currentText().strip()
        # Remove prefixes like [Obj], [Exchange]
        if text.startswith("["):
            text = text.split("]", 1)[-1].strip()
        # Get ID before " - "
        if " - " in text:
            text = text.split(" - ", 1)[0].strip()
        return text

    @Slot()
    def _autodetect_range(self):
        """Auto-detect the flux range for target reaction via FVA."""
        target_id = self._get_selected_reaction(self.target_selector)
        if not target_id:
            QMessageBox.warning(self, "No Target", "Please select a target reaction first.")
            return

        model = self.appdata.project.cobra_py_model
        if target_id not in model.reactions:
            QMessageBox.warning(self, "Invalid Reaction", f"Reaction '{target_id}' not found in model.")
            return

        try:
            self.status_label.setText("Running FVA to detect range...")
            self.autodetect_btn.setEnabled(False)

            with model:
                if self.use_scenario.isChecked():
                    try:
                        self.appdata.project.load_scenario_into_model(model)
                    except Exception:
                        pass

                target_rxn = model.reactions.get_by_id(target_id)

                # Minimize
                model.objective = target_rxn
                model.objective_direction = "min"
                sol_min = model.optimize()
                min_flux = sol_min.fluxes[target_id] if sol_min.status == "optimal" else target_rxn.lower_bound

                # Maximize
                model.objective_direction = "max"
                sol_max = model.optimize()
                max_flux = sol_max.fluxes[target_id] if sol_max.status == "optimal" else target_rxn.upper_bound

            self.min_spin.setValue(min_flux)
            self.max_spin.setValue(max_flux)
            self.status_label.setText(f"Range detected: [{min_flux:.4f}, {max_flux:.4f}]")

        except Exception as e:
            QMessageBox.warning(self, "FVA Error", f"Failed to detect range: {str(e)}")
            self.status_label.setText("")
        finally:
            self.autodetect_btn.setEnabled(True)

    @Slot()
    def _run_analysis(self):
        """Start the flux response analysis."""
        target_id = self._get_selected_reaction(self.target_selector)
        product_id = self._get_selected_reaction(self.product_selector)
        biomass_id = self._get_selected_reaction(self.biomass_selector)

        if not target_id or not product_id:
            QMessageBox.warning(self, "Missing Input", "Please select both target and product reactions.")
            return

        model = self.appdata.project.cobra_py_model

        # Validate reactions
        for rxn_id, name in [(target_id, "Target"), (product_id, "Product"), (biomass_id, "Biomass")]:
            if rxn_id and rxn_id not in model.reactions:
                QMessageBox.warning(self, "Invalid Reaction", f"{name} reaction '{rxn_id}' not found in model.")
                return

        x_min = self.min_spin.value()
        x_max = self.max_spin.value()
        if x_max < x_min:
            x_min, x_max = x_max, x_min
            self.min_spin.setValue(x_min)
            self.max_spin.setValue(x_max)

        # Disable UI, enable cancel
        self.run_btn.setEnabled(False)
        self.cancel_btn.setEnabled(True)
        self.autodetect_btn.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, self.steps_spin.value())
        self.progress_bar.setValue(0)
        self.status_label.setText("Starting analysis...")

        # Create model copy
        model_copy = model.copy()

        # Start worker thread
        self.worker_thread = FluxResponseWorkerThread(
            model_copy,
            self.appdata,
            target_id,
            product_id,
            biomass_id,
            x_min,
            x_max,
            self.steps_spin.value(),
            self.use_scenario.isChecked(),
        )
        self.worker_thread.progress_update.connect(self._on_progress)
        self.worker_thread.result_ready.connect(self._on_results)
        self.worker_thread.error_occurred.connect(self._on_error)
        self.worker_thread.finished.connect(self._on_finished)
        self.worker_thread.start()

    @Slot()
    def _cancel_analysis(self):
        """Cancel the running analysis."""
        if self.worker_thread and self.worker_thread.isRunning():
            self.worker_thread.request_cancel()
            self.status_label.setText("Cancelling...")
            self.cancel_btn.setEnabled(False)

    @Slot(int, int)
    def _on_progress(self, current: int, total: int):
        """Update progress bar."""
        self.progress_bar.setValue(current)
        self.status_label.setText(f"Scanning point {current}/{total}...")

    @Slot(object)
    def _on_results(self, data: dict):
        """Handle analysis results."""
        self.last_results = data

        scan_results = data["scan_results"]

        # Populate table
        self.results_table.setRowCount(len(scan_results))
        optimal_point = data["optimal_point"]

        for row, result in enumerate(scan_results):
            is_optimal = (
                optimal_point
                and abs(result["target_flux"] - optimal_point["target_flux"]) < 1e-9
                and result["status"] == "optimal"
            )

            # Target flux
            target_item = QTableWidgetItem(f"{result['target_flux']:.4f}")
            target_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.results_table.setItem(row, 0, target_item)

            # Product flux
            if result["status"] == "optimal":
                product_item = QTableWidgetItem(f"{result['product_flux']:.4f}")
                if is_optimal:
                    product_item.setText(f"{result['product_flux']:.4f} ★")
                    product_item.setBackground(Qt.yellow)
            else:
                product_item = QTableWidgetItem("N/A")
            product_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.results_table.setItem(row, 1, product_item)

            # Biomass
            if result["status"] == "optimal":
                biomass_item = QTableWidgetItem(f"{result['biomass_flux']:.4f}")
            else:
                biomass_item = QTableWidgetItem("N/A")
            biomass_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.results_table.setItem(row, 2, biomass_item)

            # Status
            status_item = QTableWidgetItem(result["status"])
            if result["status"] != "optimal":
                status_item.setForeground(Qt.red)
            self.results_table.setItem(row, 3, status_item)

            # Highlight optimal row
            if is_optimal:
                for col in range(4):
                    item = self.results_table.item(row, col)
                    if item:
                        item.setBackground(Qt.yellow)

        # Update plot
        if MATPLOTLIB_AVAILABLE:
            self._update_plot(data)

        # Update summary
        wt_product = data["wt_product"]
        summary_parts = [
            f"Feasible: {data['n_feasible']}/{data['n_steps']} points",
        ]

        if optimal_point:
            opt_prod = optimal_point["product_flux"]
            improvement = opt_prod - wt_product
            if wt_product != 0:
                fold = opt_prod / wt_product
                summary_parts.append(
                    f"Optimal: Target={optimal_point['target_flux']:.3f}, "
                    f"Product={opt_prod:.3f} (vs WT: {wt_product:.3f}, "
                    f"+{improvement:.3f}, {fold:.2f}x)"
                )
            else:
                summary_parts.append(
                    f"Optimal: Target={optimal_point['target_flux']:.3f}, "
                    f"Product={opt_prod:.3f} (WT: {wt_product:.3f})"
                )

        self.summary_label.setText(" | ".join(summary_parts))
        self.status_label.setText("Analysis complete!")

    def _update_plot(self, data: dict):
        """Update the plot showing Target vs Product flux."""
        self.figure.clear()

        scan_results = data["scan_results"]
        feasible = [r for r in scan_results if r["status"] == "optimal"]

        if not feasible:
            ax = self.figure.add_subplot(111)
            ax.text(0.5, 0.5, "No feasible solutions found", ha="center", va="center")
            self.canvas.draw()
            return

        target_fluxes = np.array([r["target_flux"] for r in feasible])
        product_fluxes = np.array([r["product_flux"] for r in feasible])

        optimal = data["optimal_point"]

        # Single plot: Target vs Product (similar to reference image)
        ax = self.figure.add_subplot(111)

        # Plot the line with a nice color (similar to the reference image)
        ax.plot(
            target_fluxes,
            product_fluxes,
            "-",
            linewidth=2.5,
            color="#6B4C6E",  # Purple/maroon color similar to reference
            marker="o",
            markersize=4,
            markerfacecolor="#6B4C6E",
            markeredgecolor="#6B4C6E",
        )

        # Mark optimal point with a star
        if optimal:
            ax.plot(
                optimal["target_flux"],
                optimal["product_flux"],
                "*",
                markersize=15,
                color="gold",
                markeredgecolor="black",
                markeredgewidth=0.8,
                label=f"Optimal ({optimal['product_flux']:.2f})",
                zorder=5,
            )
            ax.legend(fontsize=9, loc="best")

        # Labels with units (mmol/gDCW/h)
        ax.set_xlabel(f"{data['target_reaction']} flux (mmol/gDCW/h)", fontsize=11)
        ax.set_ylabel(f"{data['product_reaction']} flux (mmol/gDCW/h)", fontsize=11)

        # Light grid
        ax.grid(True, alpha=0.3, linestyle="-", linewidth=0.5)

        # Set axis to start from 0 if all values are positive
        if np.min(target_fluxes) >= 0:
            ax.set_xlim(left=0)
        if np.min(product_fluxes) >= 0:
            ax.set_ylim(bottom=0)

        self.figure.tight_layout()
        self.canvas.draw()

    @Slot(str)
    def _on_error(self, error: str):
        """Handle analysis error."""
        QMessageBox.warning(self, "Analysis Error", error)
        self.status_label.setText(f"Error: {error[:80]}...")

    @Slot()
    def _on_finished(self):
        """Handle worker thread completion."""
        self.run_btn.setEnabled(True)
        self.cancel_btn.setEnabled(False)
        self.autodetect_btn.setEnabled(True)
        self.progress_bar.setVisible(False)

    @Slot()
    def _export_csv(self):
        """Export results to CSV."""
        if not self.last_results:
            QMessageBox.warning(self, "No Results", "No results to export. Run analysis first.")
            return

        filename, _ = QFileDialog.getSaveFileName(
            self, "Export to CSV", self.appdata.work_directory, "CSV files (*.csv)"
        )

        if not filename:
            return

        try:
            if not filename.endswith(".csv"):
                filename += ".csv"

            with open(filename, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(
                    [
                        "Target Flux",
                        "Product Flux",
                        "Biomass Flux",
                        "Status",
                        "Improvement vs WT",
                    ]
                )

                wt_product = self.last_results["wt_product"]
                for r in self.last_results["scan_results"]:
                    improvement = (r["product_flux"] - wt_product) if r["status"] == "optimal" else ""
                    writer.writerow(
                        [
                            r["target_flux"],
                            r["product_flux"] if r["status"] == "optimal" else "",
                            r["biomass_flux"] if r["status"] == "optimal" else "",
                            r["status"],
                            improvement,
                        ]
                    )

            QMessageBox.information(self, "Exported", f"Results exported to {filename}")
        except Exception as e:
            QMessageBox.warning(self, "Export Error", f"Failed to export: {str(e)}")

    @Slot()
    def _export_xlsx(self):
        """Export results to XLSX."""
        if not self.last_results:
            QMessageBox.warning(self, "No Results", "No results to export. Run analysis first.")
            return

        if not OPENPYXL_AVAILABLE:
            QMessageBox.warning(self, "Not Available", "openpyxl is required for XLSX export.")
            return

        filename, _ = QFileDialog.getSaveFileName(
            self, "Export to XLSX", self.appdata.work_directory, "Excel files (*.xlsx)"
        )

        if not filename:
            return

        try:
            wb = openpyxl.Workbook()

            # Sheet 1: Results
            ws = wb.active
            ws.title = "Scan Results"

            headers = ["Target Flux", "Product Flux", "Biomass Flux", "Status", "Improvement vs WT"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = openpyxl.styles.Font(bold=True)

            wt_product = self.last_results["wt_product"]
            optimal = self.last_results["optimal_point"]

            for row, r in enumerate(self.last_results["scan_results"], 2):
                ws.cell(row=row, column=1, value=r["target_flux"])
                ws.cell(row=row, column=2, value=r["product_flux"] if r["status"] == "optimal" else None)
                ws.cell(row=row, column=3, value=r["biomass_flux"] if r["status"] == "optimal" else None)
                ws.cell(row=row, column=4, value=r["status"])

                if r["status"] == "optimal":
                    improvement = r["product_flux"] - wt_product
                    ws.cell(row=row, column=5, value=improvement)

                    # Highlight optimal row
                    if optimal and abs(r["target_flux"] - optimal["target_flux"]) < 1e-9:
                        for col in range(1, 6):
                            ws.cell(row=row, column=col).fill = openpyxl.styles.PatternFill(
                                start_color="FFFF00", end_color="FFFF00", fill_type="solid"
                            )

            # Sheet 2: Summary
            ws_summary = wb.create_sheet("Summary")
            ws_summary.cell(row=1, column=1, value="Parameter").font = openpyxl.styles.Font(bold=True)
            ws_summary.cell(row=1, column=2, value="Value").font = openpyxl.styles.Font(bold=True)

            summary_data = [
                ("Target Reaction", self.last_results["target_reaction"]),
                ("Product Reaction", self.last_results["product_reaction"]),
                ("Biomass Reaction", self.last_results["biomass_reaction"]),
                ("X Min", self.last_results["x_min"]),
                ("X Max", self.last_results["x_max"]),
                ("Number of Steps", self.last_results["n_steps"]),
                ("Feasible Steps", self.last_results["n_feasible"]),
                ("", ""),
                ("Wild-type Biomass", self.last_results["wt_biomass"]),
                ("Wild-type Target Flux", self.last_results["wt_target"]),
                ("Wild-type Product Flux", self.last_results["wt_product"]),
            ]

            if optimal:
                summary_data.extend(
                    [
                        ("", ""),
                        ("Optimal Target Flux", optimal["target_flux"]),
                        ("Optimal Product Flux", optimal["product_flux"]),
                        ("Optimal Biomass", optimal["biomass_flux"]),
                        ("Improvement vs WT", optimal["product_flux"] - wt_product),
                    ]
                )

            for row, (param, value) in enumerate(summary_data, 2):
                ws_summary.cell(row=row, column=1, value=param)
                ws_summary.cell(row=row, column=2, value=value)

            wb.save(filename)
            QMessageBox.information(self, "Exported", f"Results exported to {filename}")
        except Exception as e:
            QMessageBox.warning(self, "Export Error", f"Failed to export: {str(e)}")

    @Slot()
    def _export_plot(self):
        """Export the plot to a file."""
        if not MATPLOTLIB_AVAILABLE or not self.last_results:
            QMessageBox.warning(self, "No Plot", "No plot to export. Run analysis first.")
            return

        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Export Plot",
            self.appdata.work_directory,
            "PNG files (*.png);;PDF files (*.pdf);;SVG files (*.svg)",
        )

        if not filename:
            return

        try:
            self.figure.savefig(filename, dpi=150, bbox_inches="tight")
            QMessageBox.information(self, "Exported", f"Plot exported to {filename}")
        except Exception as e:
            QMessageBox.warning(self, "Export Error", f"Failed to export: {str(e)}")
