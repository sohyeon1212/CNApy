"""Batch MOMA/ROOM Analysis Dialog for CNApy

This dialog performs batch MOMA (Minimization of Metabolic Adjustment) or
ROOM (Regulatory On/Off Minimization) analysis across multiple gene or
reaction knockouts.

Features:
- Batch analysis for all genes or reactions
- Scatter plot visualization comparing WT vs KO growth
- CSV/XLSX export functionality
"""

import csv

from qtpy.QtCore import Qt, QThread, Signal, Slot
from qtpy.QtWidgets import (
    QButtonGroup,
    QComboBox,
    QDialog,
    QDoubleSpinBox,
    QFileDialog,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QRadioButton,
    QSplitter,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from cnapy.appdata import AppData
from cnapy.gui_elements.filterable_combobox import FilterableComboBox
from cnapy.gui_elements.plot_customization_dialog import PlotCustomizationDialog
from cnapy.moma import has_milp_solver, linear_moma, room

# Check for openpyxl availability for XLSX export
try:
    import openpyxl

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Check for matplotlib availability for scatter plot
try:
    from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
    from matplotlib.figure import Figure

    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False


class BatchMomaRoomWorkerThread(QThread):
    """Worker thread for running batch MOMA/ROOM analysis in background."""

    progress_update = Signal(int, int)  # (current, total)
    result_ready = Signal(object)  # dict with results
    error_occurred = Signal(str)

    def __init__(
        self,
        model,
        appdata: AppData,
        analysis_type: str,  # "moma" or "room"
        target_type: str,  # "genes" or "reactions"
        threshold: float,
        biomass_reaction: str,
        target_reaction: str | None = None,
        room_delta: float = 0.03,
        room_epsilon: float = 0.001,
        template_type: str = "fba",  # "map", "fba", "pfba"
        comp_values: dict[str, tuple[float, float]] | None = None,
    ):
        """Initialize the worker thread.

        Args:
            model: COBRApy model (copy)
            appdata: Application data for scenario loading
            analysis_type: "moma" or "room"
            target_type: "genes" or "reactions"
            threshold: Growth ratio threshold for essentiality
            biomass_reaction: ID of the biomass reaction to use for growth
            target_reaction: ID of the target product reaction (optional)
            room_delta: ROOM delta parameter
            room_epsilon: ROOM epsilon parameter
            template_type: Source for reference flux ("map", "fba", "pfba")
            comp_values: Pre-computed flux values from Map (for template_type="map")
        """
        super().__init__()
        self.model = model
        self.appdata = appdata
        self.analysis_type = analysis_type
        self.target_type = target_type
        self.threshold = threshold
        self.biomass_reaction = biomass_reaction
        self.target_reaction = target_reaction
        self.room_delta = room_delta
        self.room_epsilon = room_epsilon
        self.template_type = template_type
        self.comp_values = comp_values
        self._cancel_requested = False

    def request_cancel(self):
        """Request cancellation of the analysis."""
        self._cancel_requested = True

    def run(self):
        """Execute the batch MOMA/ROOM analysis."""
        try:
            model = self.model

            # Load scenario into model
            if self.appdata and self.appdata.project:
                self.appdata.project.load_scenario_into_model(model)

            # Get reference fluxes based on template type
            if self.template_type == "map":
                # Use pre-computed comp_values from Map
                if not self.comp_values:
                    self.error_occurred.emit("No computed flux data available in Map.")
                    return
                # Extract flux values from comp_values (tuple of (value, precision))
                reference_fluxes = {rid: vals[0] for rid, vals in self.comp_values.items()}

                # Get wild-type biomass from comp_values if available
                if self.biomass_reaction in reference_fluxes:
                    wt_biomass = reference_fluxes[self.biomass_reaction]
                else:
                    # Fallback: run FBA to get biomass
                    wt_solution = model.optimize()
                    if wt_solution.status != "optimal":
                        self.error_occurred.emit("Wild-type optimization failed. Check model constraints.")
                        return
                    wt_biomass = wt_solution.fluxes[self.biomass_reaction]

                # Get wild-type target flux if specified
                wt_target = None
                if self.target_reaction:
                    if self.target_reaction in reference_fluxes:
                        wt_target = reference_fluxes[self.target_reaction]
                    else:
                        wt_target = 0.0

            elif self.template_type == "pfba":
                # Run pFBA for reference flux
                from cobra.flux_analysis import pfba

                wt_solution = pfba(model)
                if wt_solution.status != "optimal":
                    self.error_occurred.emit("pFBA optimization failed. Check model constraints.")
                    return
                wt_biomass = wt_solution.fluxes[self.biomass_reaction]
                wt_target = wt_solution.fluxes[self.target_reaction] if self.target_reaction else None
                reference_fluxes = {rxn.id: wt_solution.fluxes[rxn.id] for rxn in model.reactions}

            else:  # "fba" (default)
                # Get wild-type growth using FBA
                wt_solution = model.optimize()
                if wt_solution.status != "optimal":
                    self.error_occurred.emit("Wild-type optimization failed. Check model constraints.")
                    return
                wt_biomass = wt_solution.fluxes[self.biomass_reaction]
                wt_target = wt_solution.fluxes[self.target_reaction] if self.target_reaction else None
                reference_fluxes = {rxn.id: wt_solution.fluxes[rxn.id] for rxn in model.reactions}

            if wt_biomass < 1e-9:
                self.error_occurred.emit("Wild-type has no growth. Cannot perform analysis.")
                return

            results = []

            if self.target_type == "genes":
                targets = list(model.genes)
                total = len(targets)
                self.progress_update.emit(0, total)

                for idx, gene in enumerate(targets):
                    if self._cancel_requested:
                        return

                    ko_biomass, ko_target = self._analyze_gene_knockout(model, gene, reference_fluxes)

                    growth_ratio = ko_biomass / wt_biomass if wt_biomass > 1e-9 else 0.0
                    is_essential = growth_ratio < self.threshold

                    result = {
                        "id": gene.id,
                        "name": gene.name if gene.name else "",
                        "type": "gene",
                        "wt_biomass": wt_biomass,
                        "ko_biomass": ko_biomass,
                        "growth_ratio": growth_ratio,
                        "is_essential": is_essential,
                    }
                    if self.target_reaction:
                        result["wt_target"] = wt_target
                        result["ko_target"] = ko_target
                        result["target_ratio"] = ko_target / wt_target if wt_target and abs(wt_target) > 1e-9 else 0.0

                    results.append(result)
                    self.progress_update.emit(idx + 1, total)

            else:  # reactions
                # Filter to reactions that can be knocked out (not exchange, not essential constraints)
                targets = [r for r in model.reactions if not r.id.startswith("EX_")]
                total = len(targets)
                self.progress_update.emit(0, total)

                for idx, reaction in enumerate(targets):
                    if self._cancel_requested:
                        return

                    ko_biomass, ko_target = self._analyze_reaction_knockout(model, reaction, reference_fluxes)

                    growth_ratio = ko_biomass / wt_biomass if wt_biomass > 1e-9 else 0.0
                    is_essential = growth_ratio < self.threshold

                    result = {
                        "id": reaction.id,
                        "name": reaction.name if reaction.name else "",
                        "type": "reaction",
                        "wt_biomass": wt_biomass,
                        "ko_biomass": ko_biomass,
                        "growth_ratio": growth_ratio,
                        "is_essential": is_essential,
                    }
                    if self.target_reaction:
                        result["wt_target"] = wt_target
                        result["ko_target"] = ko_target
                        result["target_ratio"] = ko_target / wt_target if wt_target and abs(wt_target) > 1e-9 else 0.0

                    results.append(result)
                    self.progress_update.emit(idx + 1, total)

            # Sort by growth ratio (essential first)
            results.sort(key=lambda x: x["growth_ratio"])

            n_essential = sum(1 for r in results if r["is_essential"])

            self.result_ready.emit(
                {
                    "results": results,
                    "wt_biomass": wt_biomass,
                    "wt_target": wt_target,
                    "n_essential": n_essential,
                    "n_total": len(results),
                    "threshold": self.threshold,
                    "analysis_type": self.analysis_type,
                    "target_type": self.target_type,
                    "biomass_reaction": self.biomass_reaction,
                    "target_reaction": self.target_reaction,
                }
            )

        except Exception as e:
            self.error_occurred.emit(f"Analysis failed: {str(e)}")

    def _analyze_gene_knockout(self, model, gene, reference_fluxes) -> tuple[float, float | None]:
        """Analyze a single gene knockout.

        Returns:
            Tuple of (biomass_flux, target_flux). target_flux is None if no target specified.
        """
        try:
            with model:
                # Knock out the gene
                gene.knock_out()

                # Run MOMA or ROOM
                if self.analysis_type == "moma":
                    solution = linear_moma(model, reference_fluxes)
                else:
                    solution = room(model, reference_fluxes, self.room_delta, self.room_epsilon)

                if solution.status == "optimal":
                    biomass_flux = solution.fluxes[self.biomass_reaction]
                    target_flux = solution.fluxes[self.target_reaction] if self.target_reaction else None
                    return biomass_flux, target_flux
                else:
                    return 0.0, 0.0 if self.target_reaction else None
        except Exception:
            return 0.0, 0.0 if self.target_reaction else None

    def _analyze_reaction_knockout(self, model, reaction, reference_fluxes) -> tuple[float, float | None]:
        """Analyze a single reaction knockout.

        Returns:
            Tuple of (biomass_flux, target_flux). target_flux is None if no target specified.
        """
        try:
            with model:
                # Knock out the reaction
                original_bounds = reaction.bounds
                reaction.bounds = (0, 0)

                # Run MOMA or ROOM
                if self.analysis_type == "moma":
                    solution = linear_moma(model, reference_fluxes)
                else:
                    solution = room(model, reference_fluxes, self.room_delta, self.room_epsilon)

                reaction.bounds = original_bounds

                if solution.status == "optimal":
                    biomass_flux = solution.fluxes[self.biomass_reaction]
                    target_flux = solution.fluxes[self.target_reaction] if self.target_reaction else None
                    return biomass_flux, target_flux
                else:
                    return 0.0, 0.0 if self.target_reaction else None
        except Exception:
            return 0.0, 0.0 if self.target_reaction else None


class BatchMomaRoomDialog(QDialog):
    """Dialog for batch MOMA/ROOM analysis."""

    def __init__(self, appdata: AppData, central_widget=None):
        super().__init__()
        self.setWindowTitle("Batch MOMA/ROOM Analysis")
        self.setMinimumSize(1000, 700)

        self.appdata = appdata
        self.central_widget = central_widget
        self.worker_thread: BatchMomaRoomWorkerThread | None = None
        self.last_results: dict | None = None

        self._setup_ui()

    def _setup_ui(self):
        """Setup the dialog UI."""
        main_layout = QVBoxLayout()

        # Description
        desc_label = QLabel(
            "Batch MOMA/ROOM Analysis performs knockout analysis using MOMA (Minimization of "
            "Metabolic Adjustment) or ROOM (Regulatory On/Off Minimization) for all genes or reactions.\n"
            "MOMA minimizes L1 distance to wild-type flux. ROOM minimizes the number of flux changes (requires MILP solver)."
        )
        desc_label.setWordWrap(True)
        main_layout.addWidget(desc_label)

        # Parameters group
        params_group = QGroupBox("Analysis Parameters")
        params_layout = QVBoxLayout()

        # Analysis type selection
        type_layout = QHBoxLayout()
        type_layout.addWidget(QLabel("Analysis Type:"))

        self.type_group = QButtonGroup(self)
        self.moma_radio = QRadioButton("MOMA (Linear)")
        self.moma_radio.setChecked(True)
        self.moma_radio.setToolTip("Minimization of Metabolic Adjustment - minimizes L1 distance to wild-type")
        self.type_group.addButton(self.moma_radio, 0)
        type_layout.addWidget(self.moma_radio)

        self.room_radio = QRadioButton("ROOM (MILP)")
        self.room_radio.setToolTip(
            "Regulatory On/Off Minimization - minimizes number of flux changes (requires MILP solver)"
        )
        self.type_group.addButton(self.room_radio, 1)
        type_layout.addWidget(self.room_radio)

        # Check MILP solver availability
        has_milp, milp_msg = has_milp_solver()
        if not has_milp:
            self.room_radio.setEnabled(False)
            self.room_radio.setToolTip(f"ROOM requires MILP solver. {milp_msg}")

        type_layout.addStretch()
        params_layout.addLayout(type_layout)

        # Target type selection
        target_layout = QHBoxLayout()
        target_layout.addWidget(QLabel("Target:"))

        self.target_group = QButtonGroup(self)
        self.genes_radio = QRadioButton("Genes")
        self.genes_radio.setChecked(True)
        self.genes_radio.setToolTip("Analyze gene knockouts")
        self.target_group.addButton(self.genes_radio, 0)
        target_layout.addWidget(self.genes_radio)

        self.reactions_radio = QRadioButton("Reactions")
        self.reactions_radio.setToolTip("Analyze reaction knockouts (excluding exchange reactions)")
        self.target_group.addButton(self.reactions_radio, 1)
        target_layout.addWidget(self.reactions_radio)

        target_layout.addStretch()
        params_layout.addLayout(target_layout)

        # Biomass reaction selector
        biomass_layout = QHBoxLayout()
        biomass_layout.addWidget(QLabel("Biomass Reaction:"))
        self.biomass_selector = FilterableComboBox()
        self.biomass_selector.setMinimumWidth(300)
        self.biomass_selector.setToolTip("Select the reaction to use as biomass/growth indicator")
        self._populate_biomass_selector()
        biomass_layout.addWidget(self.biomass_selector)
        biomass_layout.addStretch()
        params_layout.addLayout(biomass_layout)

        # Template flux selection
        template_layout = QHBoxLayout()
        template_layout.addWidget(QLabel("Template Flux:"))
        self.template_selector = QComboBox()
        self.template_selector.addItem("Current Map (if available)", "map")
        self.template_selector.addItem("FBA", "fba")
        self.template_selector.addItem("pFBA (parsimonious)", "pfba")
        self.template_selector.setToolTip(
            "Select the source for reference/template flux:\n"
            "• Current Map: Use computed flux from Map (LAD/E-Flux2/FBA results)\n"
            "• FBA: Run fresh FBA for reference flux\n"
            "• pFBA: Run parsimonious FBA (minimizes total flux)"
        )
        self.template_selector.setMinimumWidth(300)
        template_layout.addWidget(self.template_selector)
        template_layout.addStretch()
        params_layout.addLayout(template_layout)

        # Target product selector (optional)
        target_prod_layout = QHBoxLayout()
        target_prod_layout.addWidget(QLabel("Target Product (optional):"))
        self.target_selector = FilterableComboBox()
        self.target_selector.setMinimumWidth(300)
        self.target_selector.setToolTip("Select a target product reaction to track (optional)")
        self._populate_target_selector()
        target_prod_layout.addWidget(self.target_selector)
        target_prod_layout.addStretch()
        params_layout.addLayout(target_prod_layout)

        # Threshold setting
        threshold_layout = QHBoxLayout()
        threshold_layout.addWidget(QLabel("Essentiality threshold:"))
        self.threshold_spin = QDoubleSpinBox()
        self.threshold_spin.setRange(0.0001, 1.0)
        self.threshold_spin.setValue(0.01)
        self.threshold_spin.setDecimals(4)
        self.threshold_spin.setSingleStep(0.01)
        self.threshold_spin.setToolTip("Growth ratio below which target is considered essential (default: 0.01 = 1%)")
        threshold_layout.addWidget(self.threshold_spin)

        threshold_layout.addStretch()

        # Compute button
        self.compute_btn = QPushButton("Compute")
        self.compute_btn.setStyleSheet("font-size: 14px; padding: 8px 16px; background-color: #4CAF50; color: white;")
        self.compute_btn.clicked.connect(self._run_analysis)
        threshold_layout.addWidget(self.compute_btn)

        # Cancel button
        self.cancel_btn = QPushButton("Cancel")
        self.cancel_btn.setEnabled(False)
        self.cancel_btn.clicked.connect(self._cancel_analysis)
        threshold_layout.addWidget(self.cancel_btn)

        params_layout.addLayout(threshold_layout)
        params_group.setLayout(params_layout)
        main_layout.addWidget(params_group)

        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        main_layout.addWidget(self.progress_bar)

        # Status label
        self.status_label = QLabel("")
        self.status_label.setStyleSheet("color: gray;")
        main_layout.addWidget(self.status_label)

        # Results splitter (table and plot side by side)
        splitter = QSplitter(Qt.Horizontal)

        # Results table
        table_widget = QWidget()
        table_layout = QVBoxLayout()
        table_layout.setContentsMargins(0, 0, 0, 0)

        self.results_table = QTableWidget()
        self.results_table.setColumnCount(6)
        self.results_table.setHorizontalHeaderLabels(
            ["ID", "Name", "WT Growth", "KO Growth", "Growth Ratio", "Essential"]
        )
        self.results_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.results_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.results_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.results_table.setAlternatingRowColors(True)
        table_layout.addWidget(self.results_table)

        table_widget.setLayout(table_layout)
        splitter.addWidget(table_widget)

        # Scatter plot
        if MATPLOTLIB_AVAILABLE:
            plot_widget = QWidget()
            plot_layout = QVBoxLayout()
            plot_layout.setContentsMargins(0, 0, 0, 0)

            self.figure = Figure(figsize=(5, 4), dpi=100)
            self.canvas = FigureCanvas(self.figure)
            plot_layout.addWidget(self.canvas)

            plot_btn_row = QHBoxLayout()
            self.customize_btn = QPushButton("Customize Plot")
            self.customize_btn.setEnabled(False)
            self.customize_btn.clicked.connect(self._customize_plot)
            plot_btn_row.addWidget(self.customize_btn)
            export_plot_btn = QPushButton("Export Plot...")
            export_plot_btn.clicked.connect(self._export_plot)
            plot_btn_row.addWidget(export_plot_btn)
            plot_layout.addLayout(plot_btn_row)

            plot_widget.setLayout(plot_layout)
            splitter.addWidget(plot_widget)
        else:
            # Placeholder if matplotlib not available
            no_plot_label = QLabel("Install matplotlib for scatter plot visualization")
            no_plot_label.setAlignment(Qt.AlignCenter)
            splitter.addWidget(no_plot_label)

        splitter.setSizes([500, 400])
        main_layout.addWidget(splitter)

        # Summary and export panel
        summary_layout = QHBoxLayout()

        # Summary label
        self.summary_label = QLabel("No results yet. Click 'Compute' to start analysis.")
        summary_layout.addWidget(self.summary_label)

        summary_layout.addStretch()

        # Export buttons
        export_csv_btn = QPushButton("Export CSV")
        export_csv_btn.clicked.connect(self._export_csv)
        summary_layout.addWidget(export_csv_btn)

        export_xlsx_btn = QPushButton("Export XLSX")
        export_xlsx_btn.clicked.connect(self._export_xlsx)
        if not OPENPYXL_AVAILABLE:
            export_xlsx_btn.setEnabled(False)
            export_xlsx_btn.setToolTip("openpyxl not installed. Install with: pip install openpyxl")
        summary_layout.addWidget(export_xlsx_btn)

        main_layout.addLayout(summary_layout)

        # Close button
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(self.accept)
        btn_layout.addWidget(close_btn)
        main_layout.addLayout(btn_layout)

        self.setLayout(main_layout)

        # Update template availability based on comp_values
        self._update_template_availability()

    def _update_template_availability(self):
        """Update 'Current Map' option based on comp_values availability."""
        has_comp_values = bool(self.appdata.project.comp_values)

        model = self.template_selector.model()
        item = model.item(0)  # "Current Map" is index 0

        if has_comp_values:
            item.setEnabled(True)
            n_reactions = len(self.appdata.project.comp_values)
            item.setText(f"Current Map ({n_reactions} reactions)")
        else:
            item.setEnabled(False)
            item.setText("Current Map (no data - run FBA/pFBA/LAD/E-Flux2 first)")
            # Auto-select FBA if Current Map unavailable and it was selected
            if self.template_selector.currentIndex() == 0:
                self.template_selector.setCurrentIndex(1)

    def showEvent(self, event):
        """Update template availability when dialog is shown."""
        super().showEvent(event)
        self._update_template_availability()

    def _populate_biomass_selector(self):
        """Populate the biomass reaction selector."""
        model = self.appdata.project.cobra_py_model

        # Try to find objective reactions first
        objective_reactions = []
        for rxn in model.reactions:
            if rxn.objective_coefficient != 0:
                objective_reactions.append(rxn)

        # Add objective reactions first (likely biomass)
        for rxn in objective_reactions:
            display_text = f"{rxn.id} - {rxn.name}" if rxn.name else rxn.id
            self.biomass_selector.addItem(f"[Objective] {display_text}", rxn.id)

        # Add all other reactions
        for rxn in model.reactions:
            if rxn not in objective_reactions:
                display_text = f"{rxn.id} - {rxn.name}" if rxn.name else rxn.id
                self.biomass_selector.addItem(display_text, rxn.id)

    def _populate_target_selector(self):
        """Populate the target product selector."""
        model = self.appdata.project.cobra_py_model

        # Add "None" option
        self.target_selector.addItem("(None)", None)

        # Add exchange reactions (common targets for products)
        exchange_reactions = [r for r in model.reactions if r.id.startswith("EX_")]
        if exchange_reactions:
            for rxn in exchange_reactions:
                display_text = f"{rxn.id} - {rxn.name}" if rxn.name else rxn.id
                self.target_selector.addItem(f"[Exchange] {display_text}", rxn.id)

        # Add all other reactions
        for rxn in model.reactions:
            if not rxn.id.startswith("EX_"):
                display_text = f"{rxn.id} - {rxn.name}" if rxn.name else rxn.id
                self.target_selector.addItem(display_text, rxn.id)

    @Slot()
    def _run_analysis(self):
        """Start the batch MOMA/ROOM analysis."""
        model = self.appdata.project.cobra_py_model

        target_type = "genes" if self.genes_radio.isChecked() else "reactions"

        if target_type == "genes" and len(model.genes) == 0:
            QMessageBox.warning(self, "No Genes", "The model has no gene annotations.")
            return

        if target_type == "reactions" and len(model.reactions) == 0:
            QMessageBox.warning(self, "No Reactions", "The model has no reactions.")
            return

        # Get selected biomass and target reactions
        biomass_reaction = self.biomass_selector.currentData()
        target_reaction = self.target_selector.currentData()  # Can be None

        if not biomass_reaction:
            QMessageBox.warning(self, "No Biomass", "Please select a biomass reaction.")
            return

        analysis_type = "moma" if self.moma_radio.isChecked() else "room"

        # Get template flux settings
        template_type = self.template_selector.currentData()

        # Pass comp_values if using Current Map
        comp_values = None
        if template_type == "map":
            comp_values = dict(self.appdata.project.comp_values)

        # Disable compute, enable cancel
        self.compute_btn.setEnabled(False)
        self.cancel_btn.setEnabled(True)
        self.progress_bar.setVisible(True)

        total = len(model.genes) if target_type == "genes" else len(model.reactions)
        self.progress_bar.setRange(0, total)
        self.progress_bar.setValue(0)

        template_desc = {"map": "Current Map", "fba": "FBA", "pfba": "pFBA"}.get(template_type, template_type)
        self.status_label.setText(
            f"Starting {analysis_type.upper()} analysis on {target_type} (template: {template_desc})..."
        )

        # Create a copy of the model for the worker thread
        model_copy = model.copy()

        # Start worker thread
        self.worker_thread = BatchMomaRoomWorkerThread(
            model_copy,
            self.appdata,
            analysis_type,
            target_type,
            self.threshold_spin.value(),
            biomass_reaction,
            target_reaction,
            template_type=template_type,
            comp_values=comp_values,
        )
        self.worker_thread.progress_update.connect(self._on_progress)
        self.worker_thread.result_ready.connect(self._on_results)
        self.worker_thread.error_occurred.connect(self._on_error)
        self.worker_thread.finished.connect(self._on_finished)
        self.worker_thread.start()

    @Slot()
    def _cancel_analysis(self):
        """Cancel the running analysis."""
        if self.worker_thread and self.worker_thread.isRunning():
            self.worker_thread.request_cancel()
            self.status_label.setText("Cancelling...")
            self.cancel_btn.setEnabled(False)

    @Slot(int, int)
    def _on_progress(self, current: int, total: int):
        """Update progress bar."""
        self.progress_bar.setValue(current)
        target_type = "genes" if self.genes_radio.isChecked() else "reactions"
        self.status_label.setText(f"Analyzing {target_type[:-1]} {current}/{total}...")

    @Slot(object)
    def _on_results(self, data: dict):
        """Handle analysis results."""
        self.last_results = data
        results = data["results"]
        has_target = data.get("target_reaction") is not None

        # Configure table columns based on whether target product is selected
        if has_target:
            self.results_table.setColumnCount(9)
            self.results_table.setHorizontalHeaderLabels(
                [
                    "ID",
                    "Name",
                    "WT Biomass",
                    "KO Biomass",
                    "Growth Ratio",
                    "WT Target",
                    "KO Target",
                    "Target Ratio",
                    "Essential",
                ]
            )
        else:
            self.results_table.setColumnCount(6)
            self.results_table.setHorizontalHeaderLabels(
                ["ID", "Name", "WT Biomass", "KO Biomass", "Growth Ratio", "Essential"]
            )

        self.results_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.results_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)

        # Populate table
        self.results_table.setRowCount(len(results))

        for row, result in enumerate(results):
            col = 0
            # ID
            self.results_table.setItem(row, col, QTableWidgetItem(result["id"]))
            col += 1

            # Name
            self.results_table.setItem(row, col, QTableWidgetItem(result["name"]))
            col += 1

            # WT Biomass
            item = QTableWidgetItem(f"{result['wt_biomass']:.6f}")
            item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.results_table.setItem(row, col, item)
            col += 1

            # KO Biomass
            item = QTableWidgetItem(f"{result['ko_biomass']:.6f}")
            item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.results_table.setItem(row, col, item)
            col += 1

            # Growth Ratio
            item = QTableWidgetItem(f"{result['growth_ratio'] * 100:.2f}%")
            item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.results_table.setItem(row, col, item)
            col += 1

            # Target columns (if applicable)
            if has_target:
                # WT Target
                wt_target = result.get("wt_target", 0.0)
                item = QTableWidgetItem(f"{wt_target:.6f}" if wt_target is not None else "N/A")
                item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.results_table.setItem(row, col, item)
                col += 1

                # KO Target
                ko_target = result.get("ko_target", 0.0)
                item = QTableWidgetItem(f"{ko_target:.6f}" if ko_target is not None else "N/A")
                item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.results_table.setItem(row, col, item)
                col += 1

                # Target Ratio
                target_ratio = result.get("target_ratio", 0.0)
                item = QTableWidgetItem(f"{target_ratio * 100:.2f}%")
                item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.results_table.setItem(row, col, item)
                col += 1

            # Is Essential
            essential_text = "Yes" if result["is_essential"] else "No"
            item = QTableWidgetItem(essential_text)
            item.setTextAlignment(Qt.AlignCenter)
            if result["is_essential"]:
                item.setBackground(Qt.red)
                item.setForeground(Qt.white)
            self.results_table.setItem(row, col, item)

        # Update scatter plot
        if MATPLOTLIB_AVAILABLE:
            self._update_plot(results, data)

        # Update summary
        analysis_type = data["analysis_type"].upper()
        target_type = data["target_type"]
        summary_text = (
            f"{analysis_type}: Found {data['n_essential']} essential {target_type} out of {data['n_total']} total "
            f"(WT biomass: {data['wt_biomass']:.4f}, threshold: {data['threshold'] * 100:.2f}%)"
        )
        if has_target:
            summary_text += f" | Target: {data['target_reaction']}"
        self.summary_label.setText(summary_text)
        self.status_label.setText("Analysis complete.")

    def _update_plot(self, results: list, data: dict):
        """Update the scatter plot with results."""
        self.figure.clear()
        ax = self.figure.add_subplot(111)

        wt_biomass = data["wt_biomass"]
        has_target = data.get("target_reaction") is not None

        if has_target:
            # Plot: X = KO Biomass, Y = KO Target Production
            essential_x = []
            essential_y = []
            non_essential_x = []
            non_essential_y = []

            for result in results:
                ko_biomass = result["ko_biomass"]
                ko_target = result.get("ko_target", 0.0) or 0.0
                if result["is_essential"]:
                    essential_x.append(ko_biomass)
                    essential_y.append(ko_target)
                else:
                    non_essential_x.append(ko_biomass)
                    non_essential_y.append(ko_target)

            # Plot non-essential (green)
            if non_essential_x:
                ax.scatter(non_essential_x, non_essential_y, c="green", alpha=0.6, label="Non-essential", s=30)

            # Plot essential (red)
            if essential_x:
                ax.scatter(essential_x, essential_y, c="red", alpha=0.6, label="Essential", s=30)

            # Reference lines
            wt_target = data.get("wt_target", 0.0) or 0.0

            # Vertical line at WT biomass
            ax.axvline(x=wt_biomass, color="blue", linestyle="--", alpha=0.5, label=f"WT Biomass ({wt_biomass:.4f})")

            # Horizontal line at WT target
            if wt_target != 0:
                ax.axhline(y=wt_target, color="purple", linestyle="--", alpha=0.5, label=f"WT Target ({wt_target:.4f})")

            # Essentiality threshold line (vertical)
            threshold_x = wt_biomass * self.threshold_spin.value()
            ax.axvline(
                x=threshold_x,
                color="orange",
                linestyle=":",
                alpha=0.7,
                label=f"Threshold ({self.threshold_spin.value()*100:.0f}%)",
            )

            ax.set_xlabel("KO Biomass")
            ax.set_ylabel("KO Target Production")
            ax.set_title(f"{data['analysis_type'].upper()}: Biomass vs Target Production")

            # Set axis limits
            max_x = max(wt_biomass, max(r["ko_biomass"] for r in results) if results else 0) * 1.1
            all_targets = [r.get("ko_target", 0) or 0 for r in results] + [wt_target]
            max_y = max(all_targets) * 1.1 if all_targets else 1.0
            ax.set_xlim(0, max_x)
            ax.set_ylim(min(0, min(all_targets) * 1.1) if all_targets else 0, max_y)

        else:
            # Original plot: X = WT Biomass, Y = KO Biomass
            essential_wt = []
            essential_ko = []
            non_essential_wt = []
            non_essential_ko = []

            for result in results:
                if result["is_essential"]:
                    essential_wt.append(result["wt_biomass"])
                    essential_ko.append(result["ko_biomass"])
                else:
                    non_essential_wt.append(result["wt_biomass"])
                    non_essential_ko.append(result["ko_biomass"])

            # Plot non-essential (green)
            if non_essential_wt:
                ax.scatter(non_essential_wt, non_essential_ko, c="green", alpha=0.6, label="Non-essential", s=30)

            # Plot essential (red)
            if essential_wt:
                ax.scatter(essential_wt, essential_ko, c="red", alpha=0.6, label="Essential", s=30)

            # Reference line (y = x)
            max_val = max(wt_biomass, max(r["ko_biomass"] for r in results) if results else 0) * 1.1
            ax.plot([0, max_val], [0, max_val], "k--", alpha=0.5, label="No change (y=x)")

            # Threshold line
            threshold_y = wt_biomass * self.threshold_spin.value()
            ax.axhline(
                y=threshold_y,
                color="orange",
                linestyle=":",
                alpha=0.7,
                label=f"Threshold ({self.threshold_spin.value()*100:.0f}%)",
            )

            ax.set_xlabel("Wild-type Biomass")
            ax.set_ylabel("Knockout Biomass")
            ax.set_title(f"{data['analysis_type'].upper()} Knockout Analysis")
            ax.set_xlim(0, max_val)
            ax.set_ylim(0, max_val)

        ax.legend(loc="best", fontsize=8)
        self.canvas.draw()
        self.customize_btn.setEnabled(True)

    @Slot(str)
    def _on_error(self, error: str):
        """Handle analysis error."""
        QMessageBox.warning(self, "Analysis Error", error)
        self.status_label.setText(f"Error: {error}")

    @Slot()
    def _on_finished(self):
        """Handle worker thread completion."""
        self.compute_btn.setEnabled(True)
        self.cancel_btn.setEnabled(False)
        self.progress_bar.setVisible(False)

    @Slot()
    def _export_csv(self):
        """Export results to CSV."""
        if not self.last_results:
            QMessageBox.warning(self, "No Results", "No results to export. Run analysis first.")
            return

        filename, _ = QFileDialog.getSaveFileName(
            self, "Export to CSV", self.appdata.work_directory, "CSV files (*.csv)"
        )

        if not filename:
            return

        try:
            has_target = self.last_results.get("target_reaction") is not None
            with open(filename, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                if has_target:
                    writer.writerow(
                        [
                            "ID",
                            "Name",
                            "WT Biomass",
                            "KO Biomass",
                            "Growth Ratio",
                            "WT Target",
                            "KO Target",
                            "Target Ratio",
                            "Is Essential",
                        ]
                    )
                else:
                    writer.writerow(["ID", "Name", "WT Biomass", "KO Biomass", "Growth Ratio", "Is Essential"])

                for result in self.last_results["results"]:
                    row = [
                        result["id"],
                        result["name"],
                        result["wt_biomass"],
                        result["ko_biomass"],
                        result["growth_ratio"],
                    ]
                    if has_target:
                        row.extend(
                            [
                                result.get("wt_target", ""),
                                result.get("ko_target", ""),
                                result.get("target_ratio", ""),
                            ]
                        )
                    row.append("Yes" if result["is_essential"] else "No")
                    writer.writerow(row)

            QMessageBox.information(self, "Exported", f"Results exported to {filename}")
        except Exception as e:
            QMessageBox.warning(self, "Export Error", f"Failed to export: {str(e)}")

    @Slot()
    def _export_xlsx(self):
        """Export results to XLSX."""
        if not self.last_results:
            QMessageBox.warning(self, "No Results", "No results to export. Run analysis first.")
            return

        if not OPENPYXL_AVAILABLE:
            QMessageBox.warning(self, "Not Available", "openpyxl is required for XLSX export.")
            return

        filename, _ = QFileDialog.getSaveFileName(
            self, "Export to XLSX", self.appdata.work_directory, "Excel files (*.xlsx)"
        )

        if not filename:
            return

        try:
            has_target = self.last_results.get("target_reaction") is not None
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "MOMA-ROOM Analysis"

            # Headers
            if has_target:
                headers = [
                    "ID",
                    "Name",
                    "WT Biomass",
                    "KO Biomass",
                    "Growth Ratio",
                    "WT Target",
                    "KO Target",
                    "Target Ratio",
                    "Is Essential",
                ]
            else:
                headers = ["ID", "Name", "WT Biomass", "KO Biomass", "Growth Ratio", "Is Essential"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = openpyxl.styles.Font(bold=True)

            # Data
            for row, result in enumerate(self.last_results["results"], 2):
                col = 1
                ws.cell(row=row, column=col, value=result["id"])
                col += 1
                ws.cell(row=row, column=col, value=result["name"])
                col += 1
                ws.cell(row=row, column=col, value=result["wt_biomass"])
                col += 1
                ws.cell(row=row, column=col, value=result["ko_biomass"])
                col += 1
                ws.cell(row=row, column=col, value=result["growth_ratio"])
                col += 1

                if has_target:
                    ws.cell(row=row, column=col, value=result.get("wt_target", ""))
                    col += 1
                    ws.cell(row=row, column=col, value=result.get("ko_target", ""))
                    col += 1
                    ws.cell(row=row, column=col, value=result.get("target_ratio", ""))
                    col += 1

                ws.cell(row=row, column=col, value="Yes" if result["is_essential"] else "No")

                # Highlight essential
                if result["is_essential"]:
                    for c in range(1, len(headers) + 1):
                        ws.cell(row=row, column=c).fill = openpyxl.styles.PatternFill(
                            start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
                        )

            # Add summary sheet
            summary_ws = wb.create_sheet("Summary")
            summary_ws.cell(row=1, column=1, value="Metric")
            summary_ws.cell(row=1, column=2, value="Value")
            summary_ws.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
            summary_ws.cell(row=1, column=2).font = openpyxl.styles.Font(bold=True)

            summary_data = [
                ("Analysis Type", self.last_results["analysis_type"].upper()),
                ("Target Type", self.last_results["target_type"]),
                ("Biomass Reaction", self.last_results.get("biomass_reaction", "")),
                ("Wild-type Biomass", self.last_results["wt_biomass"]),
                ("Threshold", f"{self.last_results['threshold'] * 100:.2f}%"),
                ("Total", self.last_results["n_total"]),
                ("Essential", self.last_results["n_essential"]),
                ("Non-essential", self.last_results["n_total"] - self.last_results["n_essential"]),
            ]

            if has_target:
                summary_data.insert(3, ("Target Reaction", self.last_results.get("target_reaction", "")))
                summary_data.insert(5, ("Wild-type Target", self.last_results.get("wt_target", "")))

            for row, (metric, value) in enumerate(summary_data, 2):
                summary_ws.cell(row=row, column=1, value=metric)
                summary_ws.cell(row=row, column=2, value=value)

            wb.save(filename)
            QMessageBox.information(self, "Exported", f"Results exported to {filename}")
        except Exception as e:
            QMessageBox.warning(self, "Export Error", f"Failed to export: {str(e)}")

    @Slot()
    def _customize_plot(self):
        """Open the plot customization dialog."""
        dialog = PlotCustomizationDialog(self, self.figure, self.canvas)
        dialog.exec()

    @Slot()
    def _export_plot(self):
        """Export the scatter plot to a file."""
        if not MATPLOTLIB_AVAILABLE or not self.last_results:
            QMessageBox.warning(self, "No Plot", "No plot to export. Run analysis first.")
            return

        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Export Plot",
            self.appdata.work_directory,
            "PNG files (*.png);;PDF files (*.pdf);;SVG files (*.svg)",
        )

        if not filename:
            return

        try:
            self.figure.savefig(filename, dpi=150, bbox_inches="tight")
            QMessageBox.information(self, "Exported", f"Plot exported to {filename}")
        except Exception as e:
            QMessageBox.warning(self, "Export Error", f"Failed to export: {str(e)}")
