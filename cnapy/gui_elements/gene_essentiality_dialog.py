"""Gene Essentiality Analysis Dialog for CNApy

This dialog performs single gene deletion analysis to identify essential genes
in the metabolic model.

Essential genes are those whose knockout results in zero or near-zero growth.
"""

import csv

import numpy as np
from qtpy.QtCore import Qt, QThread, Signal, Slot
from qtpy.QtWidgets import (
    QCheckBox,
    QDialog,
    QDoubleSpinBox,
    QFileDialog,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QSplitter,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from cnapy.appdata import AppData

# Check for openpyxl availability for XLSX export
try:
    import openpyxl

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class GeneEssentialityWorkerThread(QThread):
    """Worker thread for running gene essentiality analysis in background."""

    progress_update = Signal(int, int)  # (current, total)
    result_ready = Signal(object)  # dict with results
    error_occurred = Signal(str)

    def __init__(self, model, appdata: AppData, threshold: float, include_affected: bool):
        """Initialize the worker thread.

        Args:
            model: COBRApy model (copy)
            appdata: Application data for scenario loading
            threshold: Essentiality threshold (growth ratio below which gene is essential)
            include_affected: Whether to include affected reactions in results
        """
        super().__init__()
        self.model = model
        self.appdata = appdata
        self.threshold = threshold
        self.include_affected = include_affected
        self._cancel_requested = False

    def request_cancel(self):
        """Request cancellation of the analysis."""
        self._cancel_requested = True

    def run(self):
        """Execute the gene essentiality analysis."""
        try:
            from cobra.flux_analysis import single_gene_deletion

            model = self.model

            # Load scenario into model
            if self.appdata and self.appdata.project:
                self.appdata.project.load_scenario_into_model(model)

            # Get wild-type growth
            wt_solution = model.optimize()
            if wt_solution.status != "optimal":
                self.error_occurred.emit("Wild-type optimization failed. Check model constraints.")
                return

            wt_growth = wt_solution.objective_value

            if wt_growth < 1e-9:
                self.error_occurred.emit("Wild-type has no growth. Cannot determine essential genes.")
                return

            # Perform single gene deletion analysis
            total_genes = len(model.genes)
            self.progress_update.emit(0, total_genes)

            deletion_results = single_gene_deletion(model)

            if self._cancel_requested:
                return

            # Process results
            results = []
            n_essential = 0

            for idx, (gene_idx, row) in enumerate(deletion_results.iterrows()):
                if self._cancel_requested:
                    return

                # Extract gene ID from frozenset index
                gene_id = list(gene_idx)[0] if isinstance(gene_idx, frozenset) else gene_idx
                ko_growth = row["growth"] if row["growth"] is not None else 0.0

                # Handle NaN values
                if np.isnan(ko_growth):
                    ko_growth = 0.0

                # Calculate growth ratio
                growth_ratio = ko_growth / wt_growth if wt_growth > 1e-9 else 0.0
                is_essential = growth_ratio < self.threshold

                if is_essential:
                    n_essential += 1

                # Get gene name and affected reactions
                try:
                    gene = model.genes.get_by_id(gene_id)
                    gene_name = gene.name if gene.name else ""
                    affected_reactions = [r.id for r in gene.reactions] if self.include_affected else []
                except Exception:
                    gene_name = ""
                    affected_reactions = []

                results.append(
                    {
                        "gene_id": gene_id,
                        "gene_name": gene_name,
                        "ko_growth": ko_growth,
                        "growth_ratio": growth_ratio,
                        "is_essential": is_essential,
                        "affected_reactions": affected_reactions,
                    }
                )

                self.progress_update.emit(idx + 1, total_genes)

            # Sort by growth ratio (essential genes first)
            results.sort(key=lambda x: x["growth_ratio"])

            self.result_ready.emit(
                {
                    "results": results,
                    "wt_growth": wt_growth,
                    "n_essential": n_essential,
                    "n_total": len(results),
                    "threshold": self.threshold,
                }
            )

        except Exception as e:
            self.error_occurred.emit(f"Analysis failed: {str(e)}")


class GeneEssentialityDialog(QDialog):
    """Dialog for gene essentiality analysis."""

    def __init__(self, appdata: AppData, central_widget=None):
        super().__init__()
        self.setWindowTitle("Gene Essentiality Analysis")
        self.setMinimumSize(900, 600)

        self.appdata = appdata
        self.central_widget = central_widget
        self.worker_thread: GeneEssentialityWorkerThread | None = None
        self.last_results: dict | None = None

        self._setup_ui()

    def _setup_ui(self):
        """Setup the dialog UI."""
        main_layout = QVBoxLayout()

        # Description
        desc_label = QLabel(
            "Gene Essentiality Analysis identifies genes whose knockout leads to zero or near-zero growth.\n"
            "This uses single gene deletion analysis from COBRApy."
        )
        desc_label.setWordWrap(True)
        main_layout.addWidget(desc_label)

        # Parameters group
        params_group = QGroupBox("Parameters")
        params_layout = QHBoxLayout()

        # Threshold
        threshold_layout = QHBoxLayout()
        threshold_layout.addWidget(QLabel("Essentiality threshold:"))
        self.threshold_spin = QDoubleSpinBox()
        self.threshold_spin.setRange(0.0001, 1.0)
        self.threshold_spin.setValue(0.01)
        self.threshold_spin.setDecimals(4)
        self.threshold_spin.setSingleStep(0.01)
        self.threshold_spin.setToolTip("Growth ratio below which a gene is considered essential (default: 0.01 = 1%)")
        threshold_layout.addWidget(self.threshold_spin)
        params_layout.addLayout(threshold_layout)

        # Include affected reactions
        self.include_affected_cb = QCheckBox("Include affected reactions")
        self.include_affected_cb.setChecked(True)
        self.include_affected_cb.setToolTip("Include list of reactions affected by each gene knockout")
        params_layout.addWidget(self.include_affected_cb)

        params_layout.addStretch()

        # Compute button
        self.compute_btn = QPushButton("Compute")
        self.compute_btn.setStyleSheet("font-size: 14px; padding: 8px 16px; background-color: #4CAF50; color: white;")
        self.compute_btn.clicked.connect(self._run_analysis)
        params_layout.addWidget(self.compute_btn)

        # Cancel button
        self.cancel_btn = QPushButton("Cancel")
        self.cancel_btn.setEnabled(False)
        self.cancel_btn.clicked.connect(self._cancel_analysis)
        params_layout.addWidget(self.cancel_btn)

        params_group.setLayout(params_layout)
        main_layout.addWidget(params_group)

        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        main_layout.addWidget(self.progress_bar)

        # Status label
        self.status_label = QLabel("")
        self.status_label.setStyleSheet("color: gray;")
        main_layout.addWidget(self.status_label)

        # Results splitter
        splitter = QSplitter(Qt.Vertical)

        # Results table
        table_widget = QWidget()
        table_layout = QVBoxLayout()
        table_layout.setContentsMargins(0, 0, 0, 0)

        self.results_table = QTableWidget()
        self.results_table.setColumnCount(6)
        self.results_table.setHorizontalHeaderLabels(
            ["Gene ID", "Gene Name", "KO Growth", "Growth Ratio", "Essential", "Affected Reactions"]
        )
        self.results_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.results_table.horizontalHeader().setSectionResizeMode(5, QHeaderView.Stretch)
        self.results_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.results_table.setAlternatingRowColors(True)
        self.results_table.itemSelectionChanged.connect(self._on_selection_changed)
        table_layout.addWidget(self.results_table)

        table_widget.setLayout(table_layout)
        splitter.addWidget(table_widget)

        # Summary and export panel
        summary_widget = QWidget()
        summary_layout = QHBoxLayout()
        summary_layout.setContentsMargins(5, 5, 5, 5)

        # Summary label
        self.summary_label = QLabel("No results yet. Click 'Compute' to start analysis.")
        summary_layout.addWidget(self.summary_label)

        summary_layout.addStretch()

        # Export buttons
        export_csv_btn = QPushButton("Export CSV")
        export_csv_btn.clicked.connect(self._export_csv)
        summary_layout.addWidget(export_csv_btn)

        export_xlsx_btn = QPushButton("Export XLSX")
        export_xlsx_btn.clicked.connect(self._export_xlsx)
        if not OPENPYXL_AVAILABLE:
            export_xlsx_btn.setEnabled(False)
            export_xlsx_btn.setToolTip("openpyxl not installed. Install with: pip install openpyxl")
        summary_layout.addWidget(export_xlsx_btn)

        # Highlight on map button
        self.highlight_btn = QPushButton("Highlight on Map")
        self.highlight_btn.setEnabled(False)
        self.highlight_btn.clicked.connect(self._highlight_on_map)
        self.highlight_btn.setToolTip("Highlight affected reactions of selected gene on the map")
        summary_layout.addWidget(self.highlight_btn)

        summary_widget.setLayout(summary_layout)
        summary_widget.setMaximumHeight(50)
        splitter.addWidget(summary_widget)

        splitter.setSizes([500, 50])
        main_layout.addWidget(splitter)

        # Close button
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(self.accept)
        btn_layout.addWidget(close_btn)
        main_layout.addLayout(btn_layout)

        self.setLayout(main_layout)

    @Slot()
    def _run_analysis(self):
        """Start the gene essentiality analysis."""
        model = self.appdata.project.cobra_py_model

        if len(model.genes) == 0:
            QMessageBox.warning(self, "No Genes", "The model has no gene annotations.")
            return

        # Disable compute, enable cancel
        self.compute_btn.setEnabled(False)
        self.cancel_btn.setEnabled(True)
        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, len(model.genes))
        self.progress_bar.setValue(0)
        self.status_label.setText("Starting analysis...")

        # Create a copy of the model for the worker thread
        model_copy = model.copy()

        # Start worker thread
        self.worker_thread = GeneEssentialityWorkerThread(
            model_copy, self.appdata, self.threshold_spin.value(), self.include_affected_cb.isChecked()
        )
        self.worker_thread.progress_update.connect(self._on_progress)
        self.worker_thread.result_ready.connect(self._on_results)
        self.worker_thread.error_occurred.connect(self._on_error)
        self.worker_thread.finished.connect(self._on_finished)
        self.worker_thread.start()

    @Slot()
    def _cancel_analysis(self):
        """Cancel the running analysis."""
        if self.worker_thread and self.worker_thread.isRunning():
            self.worker_thread.request_cancel()
            self.status_label.setText("Cancelling...")
            self.cancel_btn.setEnabled(False)

    @Slot(int, int)
    def _on_progress(self, current: int, total: int):
        """Update progress bar."""
        self.progress_bar.setValue(current)
        self.status_label.setText(f"Analyzing gene {current}/{total}...")

    @Slot(object)
    def _on_results(self, data: dict):
        """Handle analysis results."""
        self.last_results = data
        results = data["results"]

        # Populate table
        self.results_table.setRowCount(len(results))

        for row, result in enumerate(results):
            # Gene ID
            self.results_table.setItem(row, 0, QTableWidgetItem(result["gene_id"]))

            # Gene Name
            self.results_table.setItem(row, 1, QTableWidgetItem(result["gene_name"]))

            # KO Growth
            item = QTableWidgetItem(f"{result['ko_growth']:.6f}")
            item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.results_table.setItem(row, 2, item)

            # Growth Ratio
            item = QTableWidgetItem(f"{result['growth_ratio'] * 100:.2f}%")
            item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.results_table.setItem(row, 3, item)

            # Is Essential
            essential_text = "Yes" if result["is_essential"] else "No"
            item = QTableWidgetItem(essential_text)
            item.setTextAlignment(Qt.AlignCenter)
            if result["is_essential"]:
                item.setBackground(Qt.red)
                item.setForeground(Qt.white)
            self.results_table.setItem(row, 4, item)

            # Affected Reactions
            affected_str = "; ".join(result["affected_reactions"]) if result["affected_reactions"] else ""
            self.results_table.setItem(row, 5, QTableWidgetItem(affected_str))

        # Update summary
        self.summary_label.setText(
            f"Found {data['n_essential']} essential genes out of {data['n_total']} total "
            f"(WT growth: {data['wt_growth']:.4f}, threshold: {data['threshold'] * 100:.2f}%)"
        )
        self.status_label.setText("Analysis complete.")

    @Slot(str)
    def _on_error(self, error: str):
        """Handle analysis error."""
        QMessageBox.warning(self, "Analysis Error", error)
        self.status_label.setText(f"Error: {error}")

    @Slot()
    def _on_finished(self):
        """Handle worker thread completion."""
        self.compute_btn.setEnabled(True)
        self.cancel_btn.setEnabled(False)
        self.progress_bar.setVisible(False)

    @Slot()
    def _on_selection_changed(self):
        """Handle table selection change."""
        selected = self.results_table.selectedItems()
        self.highlight_btn.setEnabled(len(selected) > 0)

    @Slot()
    def _highlight_on_map(self):
        """Highlight affected reactions of selected gene on the map."""
        selected_rows = self.results_table.selectionModel().selectedRows()
        if not selected_rows or not self.last_results:
            return

        # Collect affected reactions from all selected genes
        affected_rxns = set()
        for index in selected_rows:
            row = index.row()
            result = self.last_results["results"][row]
            affected_rxns.update(result["affected_reactions"])

        if not affected_rxns:
            QMessageBox.information(self, "No Reactions", "Selected gene(s) have no affected reactions.")
            return

        # Clear existing comp_values and set affected reactions
        self.appdata.project.comp_values.clear()

        for rxn_id in affected_rxns:
            # Mark affected reactions with a special value to highlight them
            self.appdata.project.comp_values[rxn_id] = (0.0, 0.0)

        self.appdata.project.comp_values_type = 0

        # Update the map view
        if self.central_widget:
            self.central_widget.update()

        self.status_label.setText(f"Highlighted {len(affected_rxns)} affected reactions on the map.")

    @Slot()
    def _export_csv(self):
        """Export results to CSV."""
        if not self.last_results:
            QMessageBox.warning(self, "No Results", "No results to export. Run analysis first.")
            return

        filename, _ = QFileDialog.getSaveFileName(
            self, "Export to CSV", self.appdata.work_directory, "CSV files (*.csv)"
        )

        if not filename:
            return

        try:
            with open(filename, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(
                    ["Gene ID", "Gene Name", "KO Growth", "Growth Ratio", "Is Essential", "Affected Reactions"]
                )

                for result in self.last_results["results"]:
                    writer.writerow(
                        [
                            result["gene_id"],
                            result["gene_name"],
                            result["ko_growth"],
                            result["growth_ratio"],
                            "Yes" if result["is_essential"] else "No",
                            "; ".join(result["affected_reactions"]),
                        ]
                    )

            QMessageBox.information(self, "Exported", f"Results exported to {filename}")
        except Exception as e:
            QMessageBox.warning(self, "Export Error", f"Failed to export: {str(e)}")

    @Slot()
    def _export_xlsx(self):
        """Export results to XLSX."""
        if not self.last_results:
            QMessageBox.warning(self, "No Results", "No results to export. Run analysis first.")
            return

        if not OPENPYXL_AVAILABLE:
            QMessageBox.warning(self, "Not Available", "openpyxl is required for XLSX export.")
            return

        filename, _ = QFileDialog.getSaveFileName(
            self, "Export to XLSX", self.appdata.work_directory, "Excel files (*.xlsx)"
        )

        if not filename:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Gene Essentiality"

            # Headers
            headers = ["Gene ID", "Gene Name", "KO Growth", "Growth Ratio", "Is Essential", "Affected Reactions"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = openpyxl.styles.Font(bold=True)

            # Data
            for row, result in enumerate(self.last_results["results"], 2):
                ws.cell(row=row, column=1, value=result["gene_id"])
                ws.cell(row=row, column=2, value=result["gene_name"])
                ws.cell(row=row, column=3, value=result["ko_growth"])
                ws.cell(row=row, column=4, value=result["growth_ratio"])
                ws.cell(row=row, column=5, value="Yes" if result["is_essential"] else "No")
                ws.cell(row=row, column=6, value="; ".join(result["affected_reactions"]))

                # Highlight essential genes
                if result["is_essential"]:
                    for col in range(1, 7):
                        ws.cell(row=row, column=col).fill = openpyxl.styles.PatternFill(
                            start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
                        )

            # Add summary sheet
            summary_ws = wb.create_sheet("Summary")
            summary_ws.cell(row=1, column=1, value="Metric")
            summary_ws.cell(row=1, column=2, value="Value")
            summary_ws.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
            summary_ws.cell(row=1, column=2).font = openpyxl.styles.Font(bold=True)

            summary_data = [
                ("Wild-type Growth", self.last_results["wt_growth"]),
                ("Threshold", f"{self.last_results['threshold'] * 100:.2f}%"),
                ("Total Genes", self.last_results["n_total"]),
                ("Essential Genes", self.last_results["n_essential"]),
                ("Non-essential Genes", self.last_results["n_total"] - self.last_results["n_essential"]),
            ]

            for row, (metric, value) in enumerate(summary_data, 2):
                summary_ws.cell(row=row, column=1, value=metric)
                summary_ws.cell(row=row, column=2, value=value)

            # Auto-adjust column widths
            for ws in wb.worksheets:
                for column_cells in ws.columns:
                    length = max(len(str(cell.value) if cell.value else "") for cell in column_cells)
                    ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)

            wb.save(filename)
            QMessageBox.information(self, "Exported", f"Results exported to {filename}")
        except Exception as e:
            QMessageBox.warning(self, "Export Error", f"Failed to export: {str(e)}")

    def closeEvent(self, event):
        """Handle dialog close."""
        if self.worker_thread and self.worker_thread.isRunning():
            self.worker_thread.request_cancel()
            self.worker_thread.wait(2000)
        event.accept()
