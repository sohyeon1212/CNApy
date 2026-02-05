"""FVSEOF (Flux Variability SEOF) Dialog for CNApy

This dialog performs FVSEOF analysis to identify reactions whose flux ranges
correlate with target product flux while maintaining minimum growth constraints.

FVSEOF differs from FSEOF by performing Flux Variability Analysis (FVA) at each
target flux step, allowing identification of reactions based on their min/max flux
ranges rather than single point FBA solutions.

Algorithm:
1. Find target reaction's min/max flux range via FVA
2. Scan target flux from min to max
3. At each point:
   a. Fix target flux
   b. Run FBA to find max biomass
   c. Constrain biomass to 95% (configurable) of max
   d. Run FVA on all reactions
   e. Calculate avg flux and solution range for each reaction
4. Calculate linear regressions (slope, R, p-value) for each reaction
5. Identify up-regulation candidates (R > 0.9) and down-regulation candidates (R < -0.9)
"""

import csv
import time

import numpy as np
from qtpy.QtCore import Qt, QThread, Signal, Slot
from qtpy.QtWidgets import (
    QDialog,
    QDoubleSpinBox,
    QFileDialog,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QSpinBox,
    QSplitter,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)
from scipy import stats

from cnapy.appdata import AppData
from cnapy.gui_elements.filterable_combobox import FilterableComboBox

# Check for openpyxl availability for XLSX export
try:
    import openpyxl

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Check for matplotlib availability for plots
try:
    from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
    from matplotlib.figure import Figure

    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False


class FVSEOFWorkerThread(QThread):
    """Worker thread for running FVSEOF analysis in background."""

    progress_update = Signal(int, int, str)  # (current, total, status_message)
    stage_update = Signal(str)  # Current stage description
    time_estimate = Signal(str)  # Estimated remaining time
    result_ready = Signal(object)  # dict with results
    error_occurred = Signal(str)

    def __init__(
        self,
        model,
        appdata: AppData,
        target_reaction: str,
        biomass_reaction: str,
        n_steps: int = 10,
        biomass_fraction: float = 0.95,
    ):
        """Initialize the worker thread.

        Args:
            model: COBRApy model (copy)
            appdata: Application data for scenario loading
            target_reaction: Reaction ID to scan (target product)
            biomass_reaction: Reaction ID of the biomass/objective
            n_steps: Number of scanning steps
            biomass_fraction: Fraction of max biomass to constrain (0.0-1.0)
        """
        super().__init__()
        self.model = model
        self.appdata = appdata
        self.target_reaction = target_reaction
        self.biomass_reaction = biomass_reaction
        self.n_steps = n_steps
        self.biomass_fraction = biomass_fraction
        self._cancel_requested = False
        self._start_time = None

    def request_cancel(self):
        """Request cancellation of the analysis."""
        self._cancel_requested = True

    def _update_time_estimate(self, completed: int, total: int):
        """Update the estimated remaining time."""
        if completed == 0 or self._start_time is None:
            return

        elapsed = time.time() - self._start_time
        rate = completed / elapsed
        if rate > 0:
            remaining = (total - completed) / rate
            if remaining < 60:
                self.time_estimate.emit(f"~{int(remaining)}s remaining")
            elif remaining < 3600:
                mins = int(remaining // 60)
                secs = int(remaining % 60)
                self.time_estimate.emit(f"~{mins}m {secs}s remaining")
            else:
                hours = int(remaining // 3600)
                mins = int((remaining % 3600) // 60)
                self.time_estimate.emit(f"~{hours}h {mins}m remaining")

    def run(self):
        """Execute the FVSEOF analysis."""
        try:
            self._start_time = time.time()
            model = self.model

            # Load scenario into model
            if self.appdata and self.appdata.project:
                self.appdata.project.load_scenario_into_model(model)

            self.stage_update.emit("Step 1: Finding target flux range...")

            # Step 1: Find target reaction's min/max flux range
            target_rxn = model.reactions.get_by_id(self.target_reaction)
            biomass_rxn = model.reactions.get_by_id(self.biomass_reaction)

            # Minimize target
            with model:
                model.objective = target_rxn
                model.objective_direction = "min"
                sol_min = model.optimize()
                if sol_min.status == "optimal":
                    target_min = sol_min.fluxes[self.target_reaction]
                else:
                    target_min = target_rxn.lower_bound

            # Maximize target
            with model:
                model.objective = target_rxn
                model.objective_direction = "max"
                sol_max = model.optimize()
                if sol_max.status == "optimal":
                    target_max = sol_max.fluxes[self.target_reaction]
                else:
                    target_max = target_rxn.upper_bound

            if abs(target_max - target_min) < 1e-9:
                self.error_occurred.emit(
                    f"Target reaction {self.target_reaction} has no flexibility "
                    f"(range: {target_min:.6f} to {target_max:.6f})."
                )
                return

            # Generate scan points
            scan_points = np.linspace(target_min, target_max, self.n_steps)
            all_reactions = [rxn.id for rxn in model.reactions]
            n_reactions = len(all_reactions)

            # Estimate total operations: n_steps * n_reactions * 2 (min + max FVA)
            total_ops = self.n_steps * n_reactions * 2
            completed_ops = 0

            # Data structure for results
            # flux_info[reaction_id] = {step_idx: {min_flux, max_flux, avg_flux, sol_range, ...}}
            flux_data = []

            self.stage_update.emit("Step 2: Scanning target flux range with FVA...")

            for step_idx, target_flux in enumerate(scan_points):
                if self._cancel_requested:
                    return

                step_data = {
                    "step": step_idx + 1,
                    "target_flux": target_flux,
                }

                # Fix target reaction at current scan point
                with model:
                    target_rxn.bounds = (target_flux, target_flux)

                    # Run FBA to find max biomass at this target flux
                    model.objective = biomass_rxn
                    model.objective_direction = "max"
                    sol_biomass = model.optimize()

                    if sol_biomass.status != "optimal":
                        # Mark this step as infeasible
                        step_data["biomass"] = None
                        step_data["status"] = "infeasible"
                        for rxn_id in all_reactions:
                            step_data[f"{rxn_id}_min"] = None
                            step_data[f"{rxn_id}_max"] = None
                            step_data[f"{rxn_id}_avg"] = None
                            step_data[f"{rxn_id}_range"] = None
                        flux_data.append(step_data)
                        completed_ops += n_reactions * 2
                        continue

                    max_biomass = sol_biomass.fluxes[self.biomass_reaction]
                    constrained_biomass = max_biomass * self.biomass_fraction
                    step_data["biomass"] = max_biomass
                    step_data["constrained_biomass"] = constrained_biomass
                    step_data["status"] = "optimal"

                    # Constrain biomass at fraction of max
                    biomass_rxn.bounds = (constrained_biomass, constrained_biomass)

                    # Run FVA on all reactions
                    for rxn_idx, rxn_id in enumerate(all_reactions):
                        if self._cancel_requested:
                            return

                        rxn = model.reactions.get_by_id(rxn_id)

                        # FVA Min
                        self.progress_update.emit(
                            completed_ops,
                            total_ops,
                            f"Step {step_idx + 1}/{self.n_steps} | Rxn {rxn_idx + 1}/{n_reactions} | {rxn_id[:20]}",
                        )

                        try:
                            model.objective = rxn
                            model.objective_direction = "min"
                            sol_fva_min = model.optimize()
                            min_flux = sol_fva_min.fluxes[rxn_id] if sol_fva_min.status == "optimal" else 0.0
                        except Exception:
                            min_flux = 0.0

                        completed_ops += 1

                        # FVA Max
                        try:
                            model.objective_direction = "max"
                            sol_fva_max = model.optimize()
                            max_flux = sol_fva_max.fluxes[rxn_id] if sol_fva_max.status == "optimal" else 0.0
                        except Exception:
                            max_flux = 0.0

                        completed_ops += 1

                        # Calculate average and solution range
                        avg_flux = (min_flux + max_flux) / 2.0
                        sol_range = abs(max_flux - min_flux)

                        step_data[f"{rxn_id}_min"] = min_flux
                        step_data[f"{rxn_id}_max"] = max_flux
                        step_data[f"{rxn_id}_avg"] = avg_flux
                        step_data[f"{rxn_id}_range"] = sol_range

                        self._update_time_estimate(completed_ops, total_ops)

                flux_data.append(step_data)

            if self._cancel_requested:
                return

            self.stage_update.emit("Step 3: Calculating linear regressions...")

            # Step 4: Calculate linear regressions for each reaction
            # Filter to feasible steps only
            feasible_data = [d for d in flux_data if d.get("status") == "optimal"]

            if len(feasible_data) < 3:
                self.error_occurred.emit(
                    f"Not enough feasible points for regression analysis. "
                    f"Only {len(feasible_data)} feasible out of {self.n_steps} steps."
                )
                return

            target_fluxes = np.array([d["target_flux"] for d in feasible_data])
            regression_results = []

            for rxn_id in all_reactions:
                if rxn_id == self.target_reaction:
                    continue  # Skip target itself

                avg_fluxes = np.array([d.get(f"{rxn_id}_avg", 0.0) or 0.0 for d in feasible_data])
                min_fluxes = np.array([d.get(f"{rxn_id}_min", 0.0) or 0.0 for d in feasible_data])
                max_fluxes = np.array([d.get(f"{rxn_id}_max", 0.0) or 0.0 for d in feasible_data])
                sol_ranges = np.array([d.get(f"{rxn_id}_range", 0.0) or 0.0 for d in feasible_data])

                # Linear regression on absolute values
                try:
                    avg_slope, avg_intercept, avg_r, avg_p, avg_stderr = stats.linregress(
                        target_fluxes, np.abs(avg_fluxes)
                    )
                except Exception:
                    avg_slope, avg_intercept, avg_r, avg_p, avg_stderr = 0, 0, 0, 1, 0

                try:
                    min_slope, min_intercept, min_r, min_p, min_stderr = stats.linregress(
                        target_fluxes, np.abs(min_fluxes)
                    )
                except Exception:
                    min_slope, min_intercept, min_r, min_p, min_stderr = 0, 0, 0, 1, 0

                try:
                    max_slope, max_intercept, max_r, max_p, max_stderr = stats.linregress(
                        target_fluxes, np.abs(max_fluxes)
                    )
                except Exception:
                    max_slope, max_intercept, max_r, max_p, max_stderr = 0, 0, 0, 1, 0

                mean_sol_range = np.mean(sol_ranges)

                regression_results.append(
                    {
                        "reaction_id": rxn_id,
                        "avg_slope": avg_slope,
                        "avg_intercept": avg_intercept,
                        "avg_r": avg_r,
                        "avg_p": avg_p,
                        "avg_stderr": avg_stderr,
                        "min_slope": min_slope,
                        "min_intercept": min_intercept,
                        "min_r": min_r,
                        "min_p": min_p,
                        "min_stderr": min_stderr,
                        "max_slope": max_slope,
                        "max_intercept": max_intercept,
                        "max_r": max_r,
                        "max_p": max_p,
                        "max_stderr": max_stderr,
                        "mean_sol_range": mean_sol_range,
                    }
                )

            # Sort by absolute avg_r (filtering will be done in UI with user-adjustable cutoffs)
            regression_results.sort(key=lambda x: abs(x["avg_r"]), reverse=True)

            elapsed_time = time.time() - self._start_time

            self.result_ready.emit(
                {
                    "flux_data": flux_data,
                    "feasible_data": feasible_data,
                    "regression_results": regression_results,
                    "target_min": target_min,
                    "target_max": target_max,
                    "target_reaction": self.target_reaction,
                    "biomass_reaction": self.biomass_reaction,
                    "biomass_fraction": self.biomass_fraction,
                    "n_steps": self.n_steps,
                    "n_feasible": len(feasible_data),
                    "n_reactions": n_reactions,
                    "elapsed_time": elapsed_time,
                }
            )

        except Exception as e:
            import traceback

            self.error_occurred.emit(f"FVSEOF analysis failed: {str(e)}\n{traceback.format_exc()}")


class FVSEOFDialog(QDialog):
    """Dialog for FVSEOF (Flux Variability SEOF) analysis."""

    def __init__(self, appdata: AppData, central_widget=None):
        super().__init__()
        self.setWindowTitle("FVSEOF Analysis (Flux Variability SEOF)")
        self.setMinimumSize(1200, 800)

        self.appdata = appdata
        self.central_widget = central_widget
        self.worker_thread: FVSEOFWorkerThread | None = None
        self.last_results: dict | None = None

        self._setup_ui()

    def _setup_ui(self):
        """Setup the dialog UI."""
        main_layout = QVBoxLayout()

        # Description
        desc_label = QLabel(
            "FVSEOF performs Flux Variability Analysis at each target flux step "
            "to identify reactions whose flux ranges correlate with target production. "
            "Unlike FSEOF, FVSEOF analyzes min/max flux ranges, providing more robust "
            "identification of overexpression and knockout candidates."
        )
        desc_label.setWordWrap(True)
        main_layout.addWidget(desc_label)

        # Parameters group
        params_group = QGroupBox("Analysis Parameters")
        params_layout = QVBoxLayout()

        # Target reaction selector
        target_layout = QHBoxLayout()
        target_layout.addWidget(QLabel("Target Reaction:"))
        self.target_selector = FilterableComboBox()
        self.target_selector.setMinimumWidth(400)
        self.target_selector.setToolTip(
            "Select the target production reaction to scan (typically an exchange reaction)"
        )
        self._populate_target_selector()
        target_layout.addWidget(self.target_selector)
        target_layout.addStretch()
        params_layout.addLayout(target_layout)

        # Biomass reaction selector
        biomass_layout = QHBoxLayout()
        biomass_layout.addWidget(QLabel("Biomass Reaction:"))
        self.biomass_selector = FilterableComboBox()
        self.biomass_selector.setMinimumWidth(400)
        self.biomass_selector.setToolTip("Select the biomass/objective reaction to constrain during analysis")
        self._populate_biomass_selector()
        biomass_layout.addWidget(self.biomass_selector)
        biomass_layout.addStretch()
        params_layout.addLayout(biomass_layout)

        # Steps and biomass fraction
        config_layout = QHBoxLayout()
        config_layout.addWidget(QLabel("Number of Steps:"))
        self.nsteps_spin = QSpinBox()
        self.nsteps_spin.setRange(5, 30)
        self.nsteps_spin.setValue(10)
        self.nsteps_spin.setToolTip(
            "Number of scanning points between min and max target flux. "
            "More steps = more accuracy but longer computation time."
        )
        config_layout.addWidget(self.nsteps_spin)

        config_layout.addSpacing(30)

        config_layout.addWidget(QLabel("Biomass Fraction:"))
        self.fraction_spin = QDoubleSpinBox()
        self.fraction_spin.setRange(0.5, 1.0)
        self.fraction_spin.setValue(0.95)
        self.fraction_spin.setDecimals(2)
        self.fraction_spin.setSingleStep(0.05)
        self.fraction_spin.setToolTip("Constrain biomass to this fraction of maximum at each step (e.g., 0.95 = 95%)")
        config_layout.addWidget(self.fraction_spin)

        config_layout.addStretch()
        params_layout.addLayout(config_layout)

        # Cutoff parameters
        cutoff_layout = QHBoxLayout()
        cutoff_layout.addWidget(QLabel("R Cutoff (|R| >):"))
        self.r_cutoff_spin = QDoubleSpinBox()
        self.r_cutoff_spin.setRange(0.5, 0.99)
        self.r_cutoff_spin.setValue(0.9)
        self.r_cutoff_spin.setDecimals(2)
        self.r_cutoff_spin.setSingleStep(0.05)
        self.r_cutoff_spin.setToolTip("Minimum absolute correlation coefficient for candidate selection (default: 0.9)")
        self.r_cutoff_spin.valueChanged.connect(self._apply_cutoffs)
        cutoff_layout.addWidget(self.r_cutoff_spin)

        cutoff_layout.addSpacing(30)

        cutoff_layout.addWidget(QLabel("P-value Cutoff (<):"))
        self.p_cutoff_spin = QDoubleSpinBox()
        self.p_cutoff_spin.setRange(0.001, 0.1)
        self.p_cutoff_spin.setValue(0.05)
        self.p_cutoff_spin.setDecimals(3)
        self.p_cutoff_spin.setSingleStep(0.01)
        self.p_cutoff_spin.setToolTip("Maximum p-value for candidate selection (default: 0.05)")
        self.p_cutoff_spin.valueChanged.connect(self._apply_cutoffs)
        cutoff_layout.addWidget(self.p_cutoff_spin)

        # Compute button
        self.compute_btn = QPushButton("Compute")
        self.compute_btn.setStyleSheet("font-size: 14px; padding: 8px 16px; background-color: #4CAF50; color: white;")
        self.compute_btn.clicked.connect(self._run_analysis)
        cutoff_layout.addWidget(self.compute_btn)

        # Cancel button
        self.cancel_btn = QPushButton("Cancel")
        self.cancel_btn.setEnabled(False)
        self.cancel_btn.clicked.connect(self._cancel_analysis)
        cutoff_layout.addWidget(self.cancel_btn)

        params_layout.addLayout(cutoff_layout)
        params_group.setLayout(params_layout)
        main_layout.addWidget(params_group)

        # Progress section
        progress_group = QGroupBox("Progress")
        progress_layout = QVBoxLayout()

        self.stage_label = QLabel("Ready to start analysis.")
        progress_layout.addWidget(self.stage_label)

        progress_bar_layout = QHBoxLayout()
        self.progress_bar = QProgressBar()
        self.progress_bar.setMinimum(0)
        self.progress_bar.setMaximum(100)
        progress_bar_layout.addWidget(self.progress_bar, stretch=4)

        self.time_label = QLabel("")
        self.time_label.setMinimumWidth(150)
        progress_bar_layout.addWidget(self.time_label)
        progress_layout.addLayout(progress_bar_layout)

        self.detail_label = QLabel("")
        self.detail_label.setStyleSheet("color: gray;")
        progress_layout.addWidget(self.detail_label)

        progress_group.setLayout(progress_layout)
        main_layout.addWidget(progress_group)

        # Results splitter
        splitter = QSplitter(Qt.Horizontal)

        # Left side: Up-regulation candidates table
        up_widget = QWidget()
        up_layout = QVBoxLayout()
        up_layout.setContentsMargins(0, 0, 0, 0)
        up_label = QLabel("Up-regulation Candidates (R > 0.9, p < 0.05):")
        up_label.setStyleSheet("font-weight: bold; color: green;")
        up_layout.addWidget(up_label)

        self.up_table = QTableWidget()
        self.up_table.setColumnCount(5)
        self.up_table.setHorizontalHeaderLabels(["Reaction", "AVG R", "AVG P", "Mean Range", "Direction"])
        self.up_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.up_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.up_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.up_table.setAlternatingRowColors(True)
        up_layout.addWidget(self.up_table)
        up_widget.setLayout(up_layout)
        splitter.addWidget(up_widget)

        # Right side: Down-regulation candidates table
        down_widget = QWidget()
        down_layout = QVBoxLayout()
        down_layout.setContentsMargins(0, 0, 0, 0)
        down_label = QLabel("Down-regulation Candidates (R < -0.9, p < 0.05):")
        down_label.setStyleSheet("font-weight: bold; color: red;")
        down_layout.addWidget(down_label)

        self.down_table = QTableWidget()
        self.down_table.setColumnCount(5)
        self.down_table.setHorizontalHeaderLabels(["Reaction", "AVG R", "AVG P", "Mean Range", "Direction"])
        self.down_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.down_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.down_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.down_table.setAlternatingRowColors(True)
        down_layout.addWidget(self.down_table)
        down_widget.setLayout(down_layout)
        splitter.addWidget(down_widget)

        # Add plot if matplotlib available
        if MATPLOTLIB_AVAILABLE:
            plot_widget = QWidget()
            plot_layout = QVBoxLayout()
            plot_layout.setContentsMargins(0, 0, 0, 0)

            self.figure = Figure(figsize=(5, 4), dpi=100)
            self.canvas = FigureCanvas(self.figure)
            plot_layout.addWidget(self.canvas)

            export_plot_btn = QPushButton("Export Plot...")
            export_plot_btn.clicked.connect(self._export_plot)
            plot_layout.addWidget(export_plot_btn)

            plot_widget.setLayout(plot_layout)
            splitter.addWidget(plot_widget)

        splitter.setSizes([350, 350, 400])
        main_layout.addWidget(splitter)

        # Summary and export panel
        summary_layout = QHBoxLayout()

        self.summary_label = QLabel("No results yet. Click 'Compute' to start FVSEOF analysis.")
        summary_layout.addWidget(self.summary_label)

        summary_layout.addStretch()

        # Export buttons
        export_csv_btn = QPushButton("Export CSV")
        export_csv_btn.clicked.connect(self._export_csv)
        summary_layout.addWidget(export_csv_btn)

        export_xlsx_btn = QPushButton("Export XLSX")
        export_xlsx_btn.clicked.connect(self._export_xlsx)
        if not OPENPYXL_AVAILABLE:
            export_xlsx_btn.setEnabled(False)
            export_xlsx_btn.setToolTip("openpyxl not installed. Install with: pip install openpyxl")
        summary_layout.addWidget(export_xlsx_btn)

        main_layout.addLayout(summary_layout)

        # Close button
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(self.accept)
        btn_layout.addWidget(close_btn)
        main_layout.addLayout(btn_layout)

        self.setLayout(main_layout)

    def _populate_target_selector(self):
        """Populate the target reaction selector (prioritize exchange reactions)."""
        model = self.appdata.project.cobra_py_model

        # Add exchange reactions first (common production targets)
        exchange_reactions = [r for r in model.reactions if r.id.startswith("EX_")]
        for rxn in sorted(exchange_reactions, key=lambda r: r.id):
            display_text = f"{rxn.id} - {rxn.name}" if rxn.name else rxn.id
            self.target_selector.addItem(f"[Exchange] {display_text}", rxn.id)

        # Add all other reactions
        for rxn in sorted(model.reactions, key=lambda r: r.id):
            if not rxn.id.startswith("EX_"):
                display_text = f"{rxn.id} - {rxn.name}" if rxn.name else rxn.id
                self.target_selector.addItem(display_text, rxn.id)

    def _populate_biomass_selector(self):
        """Populate the biomass reaction selector (prioritize biomass/objective reactions)."""
        model = self.appdata.project.cobra_py_model

        # Try to find objective reactions first (reactions with non-zero objective coefficient)
        objective_reactions = []
        for rxn in model.reactions:
            if rxn.objective_coefficient != 0:
                objective_reactions.append(rxn)

        # If no objective reactions found, look for common biomass reaction patterns
        if not objective_reactions:
            biomass_patterns = ["biomass", "growth", "bio_", "bof_"]
            for rxn in model.reactions:
                rxn_id_lower = rxn.id.lower()
                rxn_name_lower = (rxn.name or "").lower()
                for pattern in biomass_patterns:
                    if pattern in rxn_id_lower or pattern in rxn_name_lower:
                        objective_reactions.append(rxn)
                        break

        # Add objective/biomass reactions first
        for rxn in objective_reactions:
            display_text = f"{rxn.id} - {rxn.name}" if rxn.name else rxn.id
            self.biomass_selector.addItem(f"[Obj] {display_text}", rxn.id)

        # Add all other reactions
        for rxn in sorted(model.reactions, key=lambda r: r.id):
            if rxn not in objective_reactions:
                display_text = f"{rxn.id} - {rxn.name}" if rxn.name else rxn.id
                self.biomass_selector.addItem(display_text, rxn.id)

    @Slot()
    def _run_analysis(self):
        """Start the FVSEOF analysis."""
        model = self.appdata.project.cobra_py_model

        target_reaction = self.target_selector.currentData()
        biomass_reaction = self.biomass_selector.currentData()

        if not target_reaction:
            QMessageBox.warning(self, "No Target", "Please select a target reaction.")
            return

        if not biomass_reaction:
            QMessageBox.warning(self, "No Biomass", "Please select a biomass reaction.")
            return

        if target_reaction == biomass_reaction:
            QMessageBox.warning(self, "Same Reaction", "Target and biomass reactions must be different.")
            return

        # Warn about computation time
        n_steps = self.nsteps_spin.value()
        n_reactions = len(model.reactions)
        total_fva = n_steps * n_reactions * 2
        reply = QMessageBox.question(
            self,
            "Confirm Analysis",
            f"FVSEOF will perform {total_fva:,} FVA optimizations "
            f"({n_steps} steps x {n_reactions} reactions x 2).\n\n"
            "This may take several minutes to hours depending on model size.\n\n"
            "Continue?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes,
        )
        if reply != QMessageBox.Yes:
            return

        # Disable compute, enable cancel
        self.compute_btn.setEnabled(False)
        self.cancel_btn.setEnabled(True)
        self.progress_bar.setValue(0)
        self.stage_label.setText("Starting FVSEOF analysis...")
        self.detail_label.setText("")
        self.time_label.setText("")

        # Create a copy of the model for the worker thread
        model_copy = model.copy()

        # Start worker thread
        self.worker_thread = FVSEOFWorkerThread(
            model_copy,
            self.appdata,
            target_reaction,
            biomass_reaction,
            n_steps,
            self.fraction_spin.value(),
        )
        self.worker_thread.progress_update.connect(self._on_progress)
        self.worker_thread.stage_update.connect(self._on_stage)
        self.worker_thread.time_estimate.connect(self._on_time_estimate)
        self.worker_thread.result_ready.connect(self._on_results)
        self.worker_thread.error_occurred.connect(self._on_error)
        self.worker_thread.finished.connect(self._on_finished)
        self.worker_thread.start()

    @Slot()
    def _cancel_analysis(self):
        """Cancel the running analysis."""
        if self.worker_thread and self.worker_thread.isRunning():
            self.worker_thread.request_cancel()
            self.stage_label.setText("Cancelling...")
            self.cancel_btn.setEnabled(False)

    @Slot(int, int, str)
    def _on_progress(self, current: int, total: int, message: str):
        """Update progress bar and detail message."""
        if total > 0:
            percent = int((current / total) * 100)
            self.progress_bar.setValue(percent)
        self.detail_label.setText(message)

    @Slot(str)
    def _on_stage(self, stage: str):
        """Update stage label."""
        self.stage_label.setText(stage)

    @Slot(str)
    def _on_time_estimate(self, estimate: str):
        """Update time estimate label."""
        self.time_label.setText(estimate)

    @Slot(object)
    def _on_results(self, data: dict):
        """Handle analysis results."""
        self.last_results = data

        # Update plot
        if MATPLOTLIB_AVAILABLE:
            self._update_plot(data)

        # Apply cutoffs to filter and display candidates
        self._apply_cutoffs()

        self.stage_label.setText("Analysis complete!")
        self.progress_bar.setValue(100)

    @Slot()
    def _apply_cutoffs(self):
        """Apply R and p-value cutoffs to filter candidates and update tables."""
        if not self.last_results:
            return

        r_cutoff = self.r_cutoff_spin.value()
        p_cutoff = self.p_cutoff_spin.value()

        regression_results = self.last_results["regression_results"]

        # Filter up-regulation candidates (R > cutoff, p < cutoff for all)
        up_candidates = [
            r
            for r in regression_results
            if r["avg_r"] > r_cutoff
            and r["min_r"] > r_cutoff
            and r["max_r"] > r_cutoff
            and r["avg_p"] < p_cutoff
            and r["min_p"] < p_cutoff
            and r["max_p"] < p_cutoff
        ]
        up_candidates.sort(key=lambda x: x["avg_r"], reverse=True)

        # Filter down-regulation candidates (R < -cutoff, p < cutoff for all)
        down_candidates = [
            r
            for r in regression_results
            if r["avg_r"] < -r_cutoff
            and r["min_r"] < -r_cutoff
            and r["max_r"] < -r_cutoff
            and r["avg_p"] < p_cutoff
            and r["min_p"] < p_cutoff
            and r["max_p"] < p_cutoff
        ]
        down_candidates.sort(key=lambda x: x["avg_r"])

        # Populate up-regulation table
        self.up_table.setRowCount(len(up_candidates))
        for row, cand in enumerate(up_candidates):
            self.up_table.setItem(row, 0, QTableWidgetItem(cand["reaction_id"]))

            r_item = QTableWidgetItem(f"{cand['avg_r']:.4f}")
            r_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            r_item.setBackground(Qt.green)
            self.up_table.setItem(row, 1, r_item)

            p_item = QTableWidgetItem(f"{cand['avg_p']:.2e}")
            p_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.up_table.setItem(row, 2, p_item)

            range_item = QTableWidgetItem(f"{cand['mean_sol_range']:.4f}")
            range_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.up_table.setItem(row, 3, range_item)

            self.up_table.setItem(row, 4, QTableWidgetItem("Overexpress"))

        # Populate down-regulation table
        self.down_table.setRowCount(len(down_candidates))
        for row, cand in enumerate(down_candidates):
            self.down_table.setItem(row, 0, QTableWidgetItem(cand["reaction_id"]))

            r_item = QTableWidgetItem(f"{cand['avg_r']:.4f}")
            r_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            r_item.setBackground(Qt.red)
            r_item.setForeground(Qt.white)
            self.down_table.setItem(row, 1, r_item)

            p_item = QTableWidgetItem(f"{cand['avg_p']:.2e}")
            p_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.down_table.setItem(row, 2, p_item)

            range_item = QTableWidgetItem(f"{cand['mean_sol_range']:.4f}")
            range_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.down_table.setItem(row, 3, range_item)

            self.down_table.setItem(row, 4, QTableWidgetItem("Knockout"))

        # Update summary
        elapsed = self.last_results.get("elapsed_time", 0)
        elapsed_str = f"{int(elapsed // 60)}m {int(elapsed % 60)}s" if elapsed >= 60 else f"{elapsed:.1f}s"

        self.summary_label.setText(
            f"Found {len(up_candidates)} up-regulation and {len(down_candidates)} down-regulation candidates "
            f"(p<{p_cutoff}, |R|>{r_cutoff}) | Feasible: {self.last_results['n_feasible']}/{self.last_results['n_steps']} steps | "
            f"Time: {elapsed_str}"
        )

    def _update_plot(self, data: dict):
        """Update the plot with results."""
        self.figure.clear()

        feasible_data = data["feasible_data"]
        if not feasible_data:
            return

        # Plot: Target vs Biomass envelope
        ax = self.figure.add_subplot(111)

        target_fluxes = [d["target_flux"] for d in feasible_data]
        biomass_values = [d["biomass"] for d in feasible_data]

        ax.plot(target_fluxes, biomass_values, "b-o", markersize=5, linewidth=2, label="Max Biomass")

        # Add constrained biomass line
        constrained_biomass = [
            d.get("constrained_biomass", d["biomass"] * data["biomass_fraction"]) for d in feasible_data
        ]
        ax.plot(
            target_fluxes,
            constrained_biomass,
            "g--",
            linewidth=1.5,
            alpha=0.7,
            label=f"Constrained ({data['biomass_fraction']*100:.0f}%)",
        )

        ax.set_xlabel(f"Target: {data['target_reaction']}")
        ax.set_ylabel(f"Biomass: {data['biomass_reaction']}")
        ax.set_title("FVSEOF Production Envelope")
        ax.legend(fontsize=8)
        ax.grid(True, alpha=0.3)

        self.figure.tight_layout()
        self.canvas.draw()

    @Slot(str)
    def _on_error(self, error: str):
        """Handle analysis error."""
        QMessageBox.warning(self, "Analysis Error", error)
        self.stage_label.setText(f"Error: {error[:100]}...")

    @Slot()
    def _on_finished(self):
        """Handle worker thread completion."""
        self.compute_btn.setEnabled(True)
        self.cancel_btn.setEnabled(False)

    @Slot()
    def _export_csv(self):
        """Export results to CSV."""
        if not self.last_results:
            QMessageBox.warning(self, "No Results", "No results to export. Run analysis first.")
            return

        filename, _ = QFileDialog.getSaveFileName(
            self, "Export to CSV", self.appdata.work_directory, "CSV files (*.csv)"
        )

        if not filename:
            return

        try:
            # Export regression results (main file)
            if not filename.endswith(".csv"):
                filename += ".csv"

            with open(filename, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(
                    [
                        "Reaction",
                        "AVG_R",
                        "AVG_P",
                        "AVG_SLOPE",
                        "MIN_R",
                        "MIN_P",
                        "MIN_SLOPE",
                        "MAX_R",
                        "MAX_P",
                        "MAX_SLOPE",
                        "MEAN_SOL_RANGE",
                    ]
                )

                for r in self.last_results["regression_results"]:
                    writer.writerow(
                        [
                            r["reaction_id"],
                            r["avg_r"],
                            r["avg_p"],
                            r["avg_slope"],
                            r["min_r"],
                            r["min_p"],
                            r["min_slope"],
                            r["max_r"],
                            r["max_p"],
                            r["max_slope"],
                            r["mean_sol_range"],
                        ]
                    )

            # Export up candidates
            up_filename = filename.replace(".csv", "_up_candidates.csv")
            with open(up_filename, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["Reaction", "AVG_R", "AVG_P", "MEAN_SOL_RANGE", "Direction"])
                for c in self.last_results["up_candidates"]:
                    writer.writerow([c["reaction_id"], c["avg_r"], c["avg_p"], c["mean_sol_range"], "Up"])

            # Export down candidates
            down_filename = filename.replace(".csv", "_down_candidates.csv")
            with open(down_filename, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["Reaction", "AVG_R", "AVG_P", "MEAN_SOL_RANGE", "Direction"])
                for c in self.last_results["down_candidates"]:
                    writer.writerow([c["reaction_id"], c["avg_r"], c["avg_p"], c["mean_sol_range"], "Down"])

            QMessageBox.information(
                self,
                "Exported",
                f"Results exported to:\n{filename}\n{up_filename}\n{down_filename}",
            )
        except Exception as e:
            QMessageBox.warning(self, "Export Error", f"Failed to export: {str(e)}")

    @Slot()
    def _export_xlsx(self):
        """Export results to XLSX."""
        if not self.last_results:
            QMessageBox.warning(self, "No Results", "No results to export. Run analysis first.")
            return

        if not OPENPYXL_AVAILABLE:
            QMessageBox.warning(self, "Not Available", "openpyxl is required for XLSX export.")
            return

        filename, _ = QFileDialog.getSaveFileName(
            self, "Export to XLSX", self.appdata.work_directory, "Excel files (*.xlsx)"
        )

        if not filename:
            return

        try:
            wb = openpyxl.Workbook()

            # Sheet 1: All regression results
            ws_all = wb.active
            ws_all.title = "All Reactions"

            headers = [
                "Reaction",
                "AVG_R",
                "AVG_P",
                "AVG_SLOPE",
                "MIN_R",
                "MIN_P",
                "MIN_SLOPE",
                "MAX_R",
                "MAX_P",
                "MAX_SLOPE",
                "MEAN_SOL_RANGE",
            ]
            for col, header in enumerate(headers, 1):
                cell = ws_all.cell(row=1, column=col, value=header)
                cell.font = openpyxl.styles.Font(bold=True)

            for row, r in enumerate(self.last_results["regression_results"], 2):
                ws_all.cell(row=row, column=1, value=r["reaction_id"])
                ws_all.cell(row=row, column=2, value=r["avg_r"])
                ws_all.cell(row=row, column=3, value=r["avg_p"])
                ws_all.cell(row=row, column=4, value=r["avg_slope"])
                ws_all.cell(row=row, column=5, value=r["min_r"])
                ws_all.cell(row=row, column=6, value=r["min_p"])
                ws_all.cell(row=row, column=7, value=r["min_slope"])
                ws_all.cell(row=row, column=8, value=r["max_r"])
                ws_all.cell(row=row, column=9, value=r["max_p"])
                ws_all.cell(row=row, column=10, value=r["max_slope"])
                ws_all.cell(row=row, column=11, value=r["mean_sol_range"])

            # Sheet 2: Up-regulation candidates
            ws_up = wb.create_sheet("Up-regulation")
            headers = ["Reaction", "AVG_R", "AVG_P", "MEAN_SOL_RANGE", "Direction"]
            for col, header in enumerate(headers, 1):
                cell = ws_up.cell(row=1, column=col, value=header)
                cell.font = openpyxl.styles.Font(bold=True)
            for row, c in enumerate(self.last_results["up_candidates"], 2):
                ws_up.cell(row=row, column=1, value=c["reaction_id"])
                ws_up.cell(row=row, column=2, value=c["avg_r"])
                ws_up.cell(row=row, column=3, value=c["avg_p"])
                ws_up.cell(row=row, column=4, value=c["mean_sol_range"])
                ws_up.cell(row=row, column=5, value="Up")
                # Green background
                for col in range(1, 6):
                    ws_up.cell(row=row, column=col).fill = openpyxl.styles.PatternFill(
                        start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"
                    )

            # Sheet 3: Down-regulation candidates
            ws_down = wb.create_sheet("Down-regulation")
            for col, header in enumerate(headers, 1):
                cell = ws_down.cell(row=1, column=col, value=header)
                cell.font = openpyxl.styles.Font(bold=True)
            for row, c in enumerate(self.last_results["down_candidates"], 2):
                ws_down.cell(row=row, column=1, value=c["reaction_id"])
                ws_down.cell(row=row, column=2, value=c["avg_r"])
                ws_down.cell(row=row, column=3, value=c["avg_p"])
                ws_down.cell(row=row, column=4, value=c["mean_sol_range"])
                ws_down.cell(row=row, column=5, value="Down")
                # Red background
                for col in range(1, 6):
                    ws_down.cell(row=row, column=col).fill = openpyxl.styles.PatternFill(
                        start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
                    )

            # Sheet 4: Summary
            ws_summary = wb.create_sheet("Summary")
            ws_summary.cell(row=1, column=1, value="Parameter").font = openpyxl.styles.Font(bold=True)
            ws_summary.cell(row=1, column=2, value="Value").font = openpyxl.styles.Font(bold=True)

            summary_data = [
                ("Target Reaction", self.last_results["target_reaction"]),
                ("Biomass Reaction", self.last_results["biomass_reaction"]),
                ("Biomass Fraction", self.last_results["biomass_fraction"]),
                ("Number of Steps", self.last_results["n_steps"]),
                ("Feasible Steps", self.last_results["n_feasible"]),
                ("Target Min Flux", self.last_results["target_min"]),
                ("Target Max Flux", self.last_results["target_max"]),
                ("Total Reactions", self.last_results["n_reactions"]),
                ("Up-regulation Candidates", len(self.last_results["up_candidates"])),
                ("Down-regulation Candidates", len(self.last_results["down_candidates"])),
                ("Elapsed Time (s)", self.last_results["elapsed_time"]),
            ]

            for row, (param, value) in enumerate(summary_data, 2):
                ws_summary.cell(row=row, column=1, value=param)
                ws_summary.cell(row=row, column=2, value=value)

            wb.save(filename)
            QMessageBox.information(self, "Exported", f"Results exported to {filename}")
        except Exception as e:
            QMessageBox.warning(self, "Export Error", f"Failed to export: {str(e)}")

    @Slot()
    def _export_plot(self):
        """Export the plot to a file."""
        if not MATPLOTLIB_AVAILABLE or not self.last_results:
            QMessageBox.warning(self, "No Plot", "No plot to export. Run analysis first.")
            return

        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Export Plot",
            self.appdata.work_directory,
            "PNG files (*.png);;PDF files (*.pdf);;SVG files (*.svg)",
        )

        if not filename:
            return

        try:
            self.figure.savefig(filename, dpi=150, bbox_inches="tight")
            QMessageBox.information(self, "Exported", f"Plot exported to {filename}")
        except Exception as e:
            QMessageBox.warning(self, "Export Error", f"Failed to export: {str(e)}")
