"""FSEOF (Flux Scanning based on Enforced Objective Flux) Dialog for CNApy

This dialog performs FSEOF analysis to identify reactions correlated with
target product flux while maintaining minimum growth constraints.

FSEOF scans a target reaction flux while maintaining a minimum objective
(e.g., biomass) flux to identify reactions whose fluxes correlate with
the target, suggesting potential overexpression/knockout targets for
metabolic engineering.

Algorithm:
1. FBA to find maximum objective value
2. Add constraint: objective >= fraction * max_objective
3. Find feasible range of target reaction (min/max FBA)
4. Scan target flux from min to max
5. At each point, fix target flux and run FBA
6. Calculate Pearson correlations between target and all reaction fluxes
"""

import csv

import numpy as np
from qtpy.QtCore import Qt, QThread, Signal, Slot
from qtpy.QtWidgets import (
    QDialog,
    QDoubleSpinBox,
    QFileDialog,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QSpinBox,
    QSplitter,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from cnapy.appdata import AppData
from cnapy.gui_elements.filterable_combobox import FilterableComboBox

# Check for openpyxl availability for XLSX export
try:
    import openpyxl

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Check for matplotlib availability for plots
try:
    from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
    from matplotlib.figure import Figure

    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False


class FSEOFWorkerThread(QThread):
    """Worker thread for running FSEOF analysis in background."""

    progress_update = Signal(int, int)  # (current, total)
    result_ready = Signal(object)  # dict with results
    error_occurred = Signal(str)

    def __init__(
        self,
        model,
        appdata: AppData,
        target_reaction: str,
        objective_reaction: str,
        objective_fraction: float,
        n_points: int,
    ):
        """Initialize the worker thread.

        Args:
            model: COBRApy model (copy)
            appdata: Application data for scenario loading
            target_reaction: Reaction ID to scan
            objective_reaction: Reaction ID of the objective (biomass)
            objective_fraction: Fraction of max objective to maintain (0.0-1.0)
            n_points: Number of scanning points
        """
        super().__init__()
        self.model = model
        self.appdata = appdata
        self.target_reaction = target_reaction
        self.objective_reaction = objective_reaction
        self.objective_fraction = objective_fraction
        self.n_points = n_points
        self._cancel_requested = False

    def request_cancel(self):
        """Request cancellation of the analysis."""
        self._cancel_requested = True

    def run(self):
        """Execute the FSEOF analysis."""
        try:
            model = self.model

            # Load scenario into model
            if self.appdata and self.appdata.project:
                self.appdata.project.load_scenario_into_model(model)

            # Step 1: Find maximum objective value via FBA
            solution = model.optimize()
            if solution.status != "optimal":
                self.error_occurred.emit("Wild-type optimization failed. Check model constraints.")
                return

            max_obj = solution.fluxes[self.objective_reaction]
            if max_obj < 1e-9:
                self.error_occurred.emit("Wild-type has no growth. Cannot perform FSEOF analysis.")
                return

            # Step 2: Add objective constraint (lower bound)
            obj_rxn = model.reactions.get_by_id(self.objective_reaction)
            min_obj = self.objective_fraction * max_obj
            original_obj_lb = obj_rxn.lower_bound
            obj_rxn.lower_bound = min_obj

            # Step 3: Find target reaction's feasible range
            target_rxn = model.reactions.get_by_id(self.target_reaction)

            # Minimize target
            with model:
                model.objective = target_rxn
                model.objective_direction = "min"
                sol_min = model.optimize()
                if sol_min.status == "optimal":
                    target_min = sol_min.fluxes[self.target_reaction]
                else:
                    target_min = target_rxn.lower_bound

            # Maximize target
            with model:
                model.objective = target_rxn
                model.objective_direction = "max"
                sol_max = model.optimize()
                if sol_max.status == "optimal":
                    target_max = sol_max.fluxes[self.target_reaction]
                else:
                    target_max = target_rxn.upper_bound

            if abs(target_max - target_min) < 1e-9:
                self.error_occurred.emit(
                    f"Target reaction {self.target_reaction} has no flexibility "
                    f"(range: {target_min:.6f} to {target_max:.6f})."
                )
                return

            # Step 4: Generate scan points
            scan_points = np.linspace(target_min, target_max, self.n_points)

            # Step 5: Scan through each point
            flux_data = []
            self.progress_update.emit(0, self.n_points)

            # Restore original objective for scanning
            model.objective = obj_rxn

            for idx, target_flux in enumerate(scan_points):
                if self._cancel_requested:
                    # Restore original constraint before returning
                    obj_rxn.lower_bound = original_obj_lb
                    return

                # Fix target reaction at current scan point
                with model:
                    original_target_lb = target_rxn.lower_bound
                    original_target_ub = target_rxn.upper_bound
                    target_rxn.bounds = (target_flux, target_flux)

                    # Run FBA
                    sol = model.optimize()

                    if sol.status == "optimal":
                        row = {
                            "target_flux": target_flux,
                            "objective_flux": sol.fluxes[self.objective_reaction],
                        }
                        for rxn in model.reactions:
                            row[rxn.id] = sol.fluxes[rxn.id]
                        flux_data.append(row)

                self.progress_update.emit(idx + 1, self.n_points)

            # Restore original constraint
            obj_rxn.lower_bound = original_obj_lb

            if not flux_data:
                self.error_occurred.emit("No feasible solutions found during scan.")
                return

            # Step 6: Calculate correlations
            # Use numpy for efficient computation
            reaction_ids = [rxn.id for rxn in model.reactions]
            target_fluxes = np.array([r["target_flux"] for r in flux_data])

            correlations = []
            for rxn_id in reaction_ids:
                if rxn_id == self.target_reaction:
                    continue  # Skip target itself (correlation = 1.0)

                rxn_fluxes = np.array([r.get(rxn_id, 0.0) for r in flux_data])

                # Calculate Pearson correlation
                if np.std(rxn_fluxes) < 1e-12 or np.std(target_fluxes) < 1e-12:
                    corr = 0.0
                else:
                    corr = np.corrcoef(target_fluxes, rxn_fluxes)[0, 1]
                    if np.isnan(corr):
                        corr = 0.0

                correlations.append(
                    {
                        "reaction_id": rxn_id,
                        "correlation": corr,
                        "abs_correlation": abs(corr),
                    }
                )

            # Sort by absolute correlation
            correlations.sort(key=lambda x: x["abs_correlation"], reverse=True)

            self.result_ready.emit(
                {
                    "flux_data": flux_data,
                    "correlations": correlations,
                    "target_min": target_min,
                    "target_max": target_max,
                    "max_objective": max_obj,
                    "min_objective": min_obj,
                    "target_reaction": self.target_reaction,
                    "objective_reaction": self.objective_reaction,
                    "objective_fraction": self.objective_fraction,
                    "n_points": self.n_points,
                    "n_feasible": len(flux_data),
                }
            )

        except Exception as e:
            self.error_occurred.emit(f"FSEOF analysis failed: {str(e)}")


class FSEOFDialog(QDialog):
    """Dialog for FSEOF (Flux Scanning based on Enforced Objective Flux) analysis."""

    def __init__(self, appdata: AppData, central_widget=None):
        super().__init__()
        self.setWindowTitle("FSEOF Analysis (Flux Scanning based on Enforced Objective Flux)")
        self.setMinimumSize(1100, 750)

        self.appdata = appdata
        self.central_widget = central_widget
        self.worker_thread: FSEOFWorkerThread | None = None
        self.last_results: dict | None = None

        self._setup_ui()

    def _setup_ui(self):
        """Setup the dialog UI."""
        main_layout = QVBoxLayout()

        # Description
        desc_label = QLabel(
            "FSEOF scans a target reaction flux while maintaining minimum growth "
            "to identify reactions correlated with production. "
            "Reactions with high positive correlation are candidates for overexpression, "
            "while those with negative correlation may be knockout targets."
        )
        desc_label.setWordWrap(True)
        main_layout.addWidget(desc_label)

        # Parameters group
        params_group = QGroupBox("Analysis Parameters")
        params_layout = QVBoxLayout()

        # Target reaction selector
        target_layout = QHBoxLayout()
        target_layout.addWidget(QLabel("Target Reaction:"))
        self.target_selector = FilterableComboBox()
        self.target_selector.setMinimumWidth(350)
        self.target_selector.setToolTip(
            "Select the target production reaction to scan (typically an exchange reaction)"
        )
        self._populate_target_selector()
        target_layout.addWidget(self.target_selector)
        target_layout.addStretch()
        params_layout.addLayout(target_layout)

        # Objective reaction selector
        objective_layout = QHBoxLayout()
        objective_layout.addWidget(QLabel("Objective Reaction:"))
        self.objective_selector = FilterableComboBox()
        self.objective_selector.setMinimumWidth(350)
        self.objective_selector.setToolTip("Select the objective reaction (typically biomass) to constrain")
        self._populate_objective_selector()
        objective_layout.addWidget(self.objective_selector)
        objective_layout.addStretch()
        params_layout.addLayout(objective_layout)

        # Objective fraction and n_points
        fraction_layout = QHBoxLayout()
        fraction_layout.addWidget(QLabel("Objective Fraction:"))
        self.fraction_spin = QDoubleSpinBox()
        self.fraction_spin.setRange(0.0, 1.0)
        self.fraction_spin.setValue(0.9)
        self.fraction_spin.setDecimals(2)
        self.fraction_spin.setSingleStep(0.05)
        self.fraction_spin.setToolTip(
            "Minimum fraction of maximum objective to maintain (e.g., 0.9 = 90% of max growth)"
        )
        fraction_layout.addWidget(self.fraction_spin)

        fraction_layout.addSpacing(30)

        fraction_layout.addWidget(QLabel("Number of Points:"))
        self.npoints_spin = QSpinBox()
        self.npoints_spin.setRange(5, 100)
        self.npoints_spin.setValue(20)
        self.npoints_spin.setToolTip("Number of scanning points between min and max target flux")
        fraction_layout.addWidget(self.npoints_spin)

        fraction_layout.addStretch()

        # Compute button
        self.compute_btn = QPushButton("Compute")
        self.compute_btn.setStyleSheet("font-size: 14px; padding: 8px 16px; background-color: #4CAF50; color: white;")
        self.compute_btn.clicked.connect(self._run_analysis)
        fraction_layout.addWidget(self.compute_btn)

        # Cancel button
        self.cancel_btn = QPushButton("Cancel")
        self.cancel_btn.setEnabled(False)
        self.cancel_btn.clicked.connect(self._cancel_analysis)
        fraction_layout.addWidget(self.cancel_btn)

        params_layout.addLayout(fraction_layout)
        params_group.setLayout(params_layout)
        main_layout.addWidget(params_group)

        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        main_layout.addWidget(self.progress_bar)

        # Status label
        self.status_label = QLabel("")
        self.status_label.setStyleSheet("color: gray;")
        main_layout.addWidget(self.status_label)

        # Results splitter (table and plot side by side)
        splitter = QSplitter(Qt.Horizontal)

        # Results table
        table_widget = QWidget()
        table_layout = QVBoxLayout()
        table_layout.setContentsMargins(0, 0, 0, 0)

        table_label = QLabel("Top Correlated Reactions:")
        table_layout.addWidget(table_label)

        self.results_table = QTableWidget()
        self.results_table.setColumnCount(3)
        self.results_table.setHorizontalHeaderLabels(["Reaction ID", "Correlation", "Direction"])
        self.results_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.results_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.results_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.results_table.setAlternatingRowColors(True)
        table_layout.addWidget(self.results_table)

        table_widget.setLayout(table_layout)
        splitter.addWidget(table_widget)

        # Plot panel
        if MATPLOTLIB_AVAILABLE:
            plot_widget = QWidget()
            plot_layout = QVBoxLayout()
            plot_layout.setContentsMargins(0, 0, 0, 0)

            self.figure = Figure(figsize=(6, 5), dpi=100)
            self.canvas = FigureCanvas(self.figure)
            plot_layout.addWidget(self.canvas)

            # Export plot button
            export_plot_btn = QPushButton("Export Plot...")
            export_plot_btn.clicked.connect(self._export_plot)
            plot_layout.addWidget(export_plot_btn)

            plot_widget.setLayout(plot_layout)
            splitter.addWidget(plot_widget)
        else:
            no_plot_label = QLabel("Install matplotlib for visualization")
            no_plot_label.setAlignment(Qt.AlignCenter)
            splitter.addWidget(no_plot_label)

        splitter.setSizes([450, 550])
        main_layout.addWidget(splitter)

        # Summary and export panel
        summary_layout = QHBoxLayout()

        # Summary label
        self.summary_label = QLabel("No results yet. Click 'Compute' to start FSEOF analysis.")
        summary_layout.addWidget(self.summary_label)

        summary_layout.addStretch()

        # Export buttons
        export_csv_btn = QPushButton("Export CSV")
        export_csv_btn.clicked.connect(self._export_csv)
        summary_layout.addWidget(export_csv_btn)

        export_xlsx_btn = QPushButton("Export XLSX")
        export_xlsx_btn.clicked.connect(self._export_xlsx)
        if not OPENPYXL_AVAILABLE:
            export_xlsx_btn.setEnabled(False)
            export_xlsx_btn.setToolTip("openpyxl not installed. Install with: pip install openpyxl")
        summary_layout.addWidget(export_xlsx_btn)

        main_layout.addLayout(summary_layout)

        # Close button
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(self.accept)
        btn_layout.addWidget(close_btn)
        main_layout.addLayout(btn_layout)

        self.setLayout(main_layout)

    def _populate_target_selector(self):
        """Populate the target reaction selector (prioritize exchange reactions)."""
        model = self.appdata.project.cobra_py_model

        # Add exchange reactions first (common production targets)
        exchange_reactions = [r for r in model.reactions if r.id.startswith("EX_")]
        for rxn in sorted(exchange_reactions, key=lambda r: r.id):
            display_text = f"{rxn.id} - {rxn.name}" if rxn.name else rxn.id
            self.target_selector.addItem(f"[Exchange] {display_text}", rxn.id)

        # Add all other reactions
        for rxn in sorted(model.reactions, key=lambda r: r.id):
            if not rxn.id.startswith("EX_"):
                display_text = f"{rxn.id} - {rxn.name}" if rxn.name else rxn.id
                self.target_selector.addItem(display_text, rxn.id)

    def _populate_objective_selector(self):
        """Populate the objective reaction selector (prioritize biomass/objective reactions)."""
        model = self.appdata.project.cobra_py_model

        # Try to find objective reactions first (reactions with non-zero objective coefficient)
        objective_reactions = []
        for rxn in model.reactions:
            if rxn.objective_coefficient != 0:
                objective_reactions.append(rxn)

        # If no objective reactions found, look for common biomass reaction patterns
        if not objective_reactions:
            biomass_patterns = ["biomass", "growth", "bio_", "bof_"]
            for rxn in model.reactions:
                rxn_id_lower = rxn.id.lower()
                rxn_name_lower = (rxn.name or "").lower()
                for pattern in biomass_patterns:
                    if pattern in rxn_id_lower or pattern in rxn_name_lower:
                        objective_reactions.append(rxn)
                        break

        # Add objective/biomass reactions first
        for rxn in objective_reactions:
            display_text = f"{rxn.id} - {rxn.name}" if rxn.name else rxn.id
            self.objective_selector.addItem(f"[Obj] {display_text}", rxn.id)

        # Add all other reactions
        for rxn in sorted(model.reactions, key=lambda r: r.id):
            if rxn not in objective_reactions:
                display_text = f"{rxn.id} - {rxn.name}" if rxn.name else rxn.id
                self.objective_selector.addItem(display_text, rxn.id)

    @Slot()
    def _run_analysis(self):
        """Start the FSEOF analysis."""
        model = self.appdata.project.cobra_py_model

        target_reaction = self.target_selector.currentData()
        objective_reaction = self.objective_selector.currentData()

        if not target_reaction:
            QMessageBox.warning(self, "No Target", "Please select a target reaction.")
            return

        if not objective_reaction:
            QMessageBox.warning(self, "No Objective", "Please select an objective reaction.")
            return

        if target_reaction == objective_reaction:
            QMessageBox.warning(self, "Same Reaction", "Target and objective reactions must be different.")
            return

        # Disable compute, enable cancel
        self.compute_btn.setEnabled(False)
        self.cancel_btn.setEnabled(True)
        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, self.npoints_spin.value())
        self.progress_bar.setValue(0)
        self.status_label.setText("Starting FSEOF analysis...")

        # Create a copy of the model for the worker thread
        model_copy = model.copy()

        # Start worker thread
        self.worker_thread = FSEOFWorkerThread(
            model_copy,
            self.appdata,
            target_reaction,
            objective_reaction,
            self.fraction_spin.value(),
            self.npoints_spin.value(),
        )
        self.worker_thread.progress_update.connect(self._on_progress)
        self.worker_thread.result_ready.connect(self._on_results)
        self.worker_thread.error_occurred.connect(self._on_error)
        self.worker_thread.finished.connect(self._on_finished)
        self.worker_thread.start()

    @Slot()
    def _cancel_analysis(self):
        """Cancel the running analysis."""
        if self.worker_thread and self.worker_thread.isRunning():
            self.worker_thread.request_cancel()
            self.status_label.setText("Cancelling...")
            self.cancel_btn.setEnabled(False)

    @Slot(int, int)
    def _on_progress(self, current: int, total: int):
        """Update progress bar."""
        self.progress_bar.setValue(current)
        self.status_label.setText(f"Scanning point {current}/{total}...")

    @Slot(object)
    def _on_results(self, data: dict):
        """Handle analysis results."""
        self.last_results = data

        correlations = data["correlations"]

        # Populate table with top 100 correlations
        display_count = min(100, len(correlations))
        self.results_table.setRowCount(display_count)

        for row, corr_data in enumerate(correlations[:display_count]):
            # Reaction ID
            rxn_id = corr_data["reaction_id"]
            self.results_table.setItem(row, 0, QTableWidgetItem(rxn_id))

            # Correlation value
            corr_val = corr_data["correlation"]
            item = QTableWidgetItem(f"{corr_val:.4f}")
            item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)

            # Color code by correlation
            if corr_val > 0.5:
                item.setBackground(Qt.green)
            elif corr_val < -0.5:
                item.setBackground(Qt.red)
                item.setForeground(Qt.white)
            self.results_table.setItem(row, 1, item)

            # Direction
            if corr_val > 0.5:
                direction = "Overexpress"
            elif corr_val < -0.5:
                direction = "Knockout"
            else:
                direction = "-"
            self.results_table.setItem(row, 2, QTableWidgetItem(direction))

        # Update plot
        if MATPLOTLIB_AVAILABLE:
            self._update_plot(data)

        # Update summary
        n_positive = sum(1 for c in correlations if c["correlation"] > 0.5)
        n_negative = sum(1 for c in correlations if c["correlation"] < -0.5)

        top_positive = correlations[0] if correlations and correlations[0]["correlation"] > 0 else None
        top_negative = next((c for c in correlations if c["correlation"] < 0), None)

        summary_parts = [
            f"Found {len(correlations)} reactions.",
            f"Strongly correlated (|r|>0.5): {n_positive} positive, {n_negative} negative.",
        ]
        if top_positive:
            summary_parts.append(f"Top positive: {top_positive['reaction_id']} ({top_positive['correlation']:.3f})")
        if top_negative:
            summary_parts.append(f"Top negative: {top_negative['reaction_id']} ({top_negative['correlation']:.3f})")

        self.summary_label.setText(" | ".join(summary_parts))
        self.status_label.setText(
            f"Analysis complete. Scanned {data['n_feasible']}/{data['n_points']} feasible points."
        )

    def _update_plot(self, data: dict):
        """Update the plots with results."""
        self.figure.clear()

        flux_data = data["flux_data"]
        correlations = data["correlations"]

        if not flux_data:
            return

        # Plot 1: Production Envelope (Target vs Objective)
        ax1 = self.figure.add_subplot(121)
        target_fluxes = [r["target_flux"] for r in flux_data]
        obj_fluxes = [r["objective_flux"] for r in flux_data]

        ax1.plot(target_fluxes, obj_fluxes, "b-o", markersize=4, linewidth=1.5)
        ax1.set_xlabel(f"Target: {data['target_reaction']}")
        ax1.set_ylabel(f"Objective: {data['objective_reaction']}")
        ax1.set_title("Production Envelope")
        ax1.grid(True, alpha=0.3)

        # Add reference lines
        ax1.axhline(
            y=data["min_objective"],
            color="orange",
            linestyle="--",
            alpha=0.7,
            label=f"Min Obj ({data['objective_fraction']*100:.0f}%)",
        )
        ax1.legend(fontsize=8)

        # Plot 2: Top Correlations Bar Chart
        ax2 = self.figure.add_subplot(122)
        top_corr = correlations[:15]  # Top 15

        if top_corr:
            rxn_ids = [c["reaction_id"] for c in top_corr]
            corr_vals = [c["correlation"] for c in top_corr]
            colors = ["green" if v > 0 else "red" for v in corr_vals]

            y_pos = range(len(rxn_ids))
            ax2.barh(y_pos, corr_vals, color=colors, alpha=0.7)
            ax2.set_yticks(y_pos)
            ax2.set_yticklabels(rxn_ids, fontsize=7)
            ax2.set_xlabel("Correlation with Target")
            ax2.set_title("Top Correlated Reactions")
            ax2.axvline(x=0, color="black", linestyle="-", linewidth=0.5)
            ax2.set_xlim(-1.1, 1.1)
            ax2.invert_yaxis()  # Top correlation at top

        self.figure.tight_layout()
        self.canvas.draw()

    @Slot(str)
    def _on_error(self, error: str):
        """Handle analysis error."""
        QMessageBox.warning(self, "Analysis Error", error)
        self.status_label.setText(f"Error: {error}")

    @Slot()
    def _on_finished(self):
        """Handle worker thread completion."""
        self.compute_btn.setEnabled(True)
        self.cancel_btn.setEnabled(False)
        self.progress_bar.setVisible(False)

    @Slot()
    def _export_csv(self):
        """Export results to CSV."""
        if not self.last_results:
            QMessageBox.warning(self, "No Results", "No results to export. Run analysis first.")
            return

        filename, _ = QFileDialog.getSaveFileName(
            self, "Export to CSV", self.appdata.work_directory, "CSV files (*.csv)"
        )

        if not filename:
            return

        try:
            # Export correlations
            corr_filename = filename
            if not corr_filename.endswith(".csv"):
                corr_filename += ".csv"

            with open(corr_filename, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["Reaction ID", "Correlation", "Abs Correlation"])

                for corr_data in self.last_results["correlations"]:
                    writer.writerow(
                        [
                            corr_data["reaction_id"],
                            corr_data["correlation"],
                            corr_data["abs_correlation"],
                        ]
                    )

            # Also export scan results
            scan_filename = corr_filename.replace(".csv", "_scan.csv")
            with open(scan_filename, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)

                # Get all reaction IDs for header
                flux_data = self.last_results["flux_data"]
                if flux_data:
                    # Build header from first row
                    first_row = flux_data[0]
                    headers = ["target_flux", "objective_flux"] + [
                        k for k in first_row.keys() if k not in ["target_flux", "objective_flux"]
                    ]
                    writer.writerow(headers)

                    for row in flux_data:
                        writer.writerow([row.get(h, "") for h in headers])

            QMessageBox.information(
                self,
                "Exported",
                f"Results exported to:\n{corr_filename}\n{scan_filename}",
            )
        except Exception as e:
            QMessageBox.warning(self, "Export Error", f"Failed to export: {str(e)}")

    @Slot()
    def _export_xlsx(self):
        """Export results to XLSX."""
        if not self.last_results:
            QMessageBox.warning(self, "No Results", "No results to export. Run analysis first.")
            return

        if not OPENPYXL_AVAILABLE:
            QMessageBox.warning(self, "Not Available", "openpyxl is required for XLSX export.")
            return

        filename, _ = QFileDialog.getSaveFileName(
            self, "Export to XLSX", self.appdata.work_directory, "Excel files (*.xlsx)"
        )

        if not filename:
            return

        try:
            wb = openpyxl.Workbook()

            # Sheet 1: Correlations
            ws_corr = wb.active
            ws_corr.title = "Correlations"

            headers = ["Reaction ID", "Correlation", "Abs Correlation"]
            for col, header in enumerate(headers, 1):
                cell = ws_corr.cell(row=1, column=col, value=header)
                cell.font = openpyxl.styles.Font(bold=True)

            for row, corr_data in enumerate(self.last_results["correlations"], 2):
                ws_corr.cell(row=row, column=1, value=corr_data["reaction_id"])
                ws_corr.cell(row=row, column=2, value=corr_data["correlation"])
                ws_corr.cell(row=row, column=3, value=corr_data["abs_correlation"])

                # Color code
                corr = corr_data["correlation"]
                if corr > 0.5:
                    fill = openpyxl.styles.PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                elif corr < -0.5:
                    fill = openpyxl.styles.PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                else:
                    fill = None

                if fill:
                    for col in range(1, 4):
                        ws_corr.cell(row=row, column=col).fill = fill

            # Sheet 2: Scan Results
            ws_scan = wb.create_sheet("Scan Results")
            flux_data = self.last_results["flux_data"]

            if flux_data:
                first_row = flux_data[0]
                headers = ["target_flux", "objective_flux"] + sorted(
                    [k for k in first_row.keys() if k not in ["target_flux", "objective_flux"]]
                )

                for col, header in enumerate(headers, 1):
                    cell = ws_scan.cell(row=1, column=col, value=header)
                    cell.font = openpyxl.styles.Font(bold=True)

                for row_idx, row_data in enumerate(flux_data, 2):
                    for col_idx, header in enumerate(headers, 1):
                        ws_scan.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))

            # Sheet 3: Summary
            ws_summary = wb.create_sheet("Summary")
            ws_summary.cell(row=1, column=1, value="Parameter")
            ws_summary.cell(row=1, column=2, value="Value")
            ws_summary.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
            ws_summary.cell(row=1, column=2).font = openpyxl.styles.Font(bold=True)

            summary_data = [
                ("Target Reaction", self.last_results["target_reaction"]),
                ("Objective Reaction", self.last_results["objective_reaction"]),
                ("Objective Fraction", self.last_results["objective_fraction"]),
                ("Number of Points", self.last_results["n_points"]),
                ("Feasible Points", self.last_results["n_feasible"]),
                ("Target Min Flux", self.last_results["target_min"]),
                ("Target Max Flux", self.last_results["target_max"]),
                ("Max Objective", self.last_results["max_objective"]),
                ("Min Objective (constrained)", self.last_results["min_objective"]),
                ("Total Reactions Analyzed", len(self.last_results["correlations"])),
            ]

            for row, (param, value) in enumerate(summary_data, 2):
                ws_summary.cell(row=row, column=1, value=param)
                ws_summary.cell(row=row, column=2, value=value)

            wb.save(filename)
            QMessageBox.information(self, "Exported", f"Results exported to {filename}")
        except Exception as e:
            QMessageBox.warning(self, "Export Error", f"Failed to export: {str(e)}")

    @Slot()
    def _export_plot(self):
        """Export the plot to a file."""
        if not MATPLOTLIB_AVAILABLE or not self.last_results:
            QMessageBox.warning(self, "No Plot", "No plot to export. Run analysis first.")
            return

        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Export Plot",
            self.appdata.work_directory,
            "PNG files (*.png);;PDF files (*.pdf);;SVG files (*.svg)",
        )

        if not filename:
            return

        try:
            self.figure.savefig(filename, dpi=150, bbox_inches="tight")
            QMessageBox.information(self, "Exported", f"Plot exported to {filename}")
        except Exception as e:
            QMessageBox.warning(self, "Export Error", f"Failed to export: {str(e)}")
